"""app.py - DCR Flask v1"""
import os, json, uuid, io, base64, datetime
from flask import (Flask, request, session, redirect, url_for,
                   jsonify, make_response, send_file, g)
import db
from utils import (compute_expected_reply, compute_duration, is_overdue,
                   format_date, get_next_doc_no, extract_rev, STATUS_COLORS,
                   REJECTED_STATUSES, invalidate_holidays_cache)

from config import IS_RENDER, SECRET_KEY, SESSION_CONFIG
from html_render import render_login, render_dashboard, render_register
from blueprints.projects import projects_bp
from blueprints.doc_types import doc_types_bp
from blueprints.columns import columns_bp
from blueprints.records import records_bp
from blueprints.analytics import analytics_bp
from blueprints.lists import lists_bp
from blueprints.summary import summary_bp
from blueprints.logos import logos_bp

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config.update(**SESSION_CONFIG)
app.register_blueprint(projects_bp)
app.register_blueprint(doc_types_bp)
app.register_blueprint(columns_bp)
app.register_blueprint(records_bp)
app.register_blueprint(analytics_bp)
app.register_blueprint(lists_bp)
app.register_blueprint(summary_bp)
app.register_blueprint(logos_bp)

# ── Auth helpers (extracted to auth.py — Step 2 refactor) ─────
from auth import current_user, require_login, require_superadmin, can_edit

# ── Pages ─────────────────────────────────────────────────────
@app.route("/ping")
def ping():
    return jsonify(status="ok", time=datetime.datetime.now().isoformat())

@app.route("/health")
def health():
    try:
        db.get_projects()
        return jsonify(status="ok", db="connected")
    except Exception as e:
        return jsonify(status="error", db=str(e)), 503

@app.route("/api/keepalive")
def keepalive():
    """Called periodically to prevent free-tier sleep."""
    return jsonify(status="ok", ts=datetime.datetime.now().isoformat())

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        data  = request.get_json(silent=True) or {}
        uname = data.get("username","").strip().lower()
        pw    = data.get("password","")
        if db.verify_pw(uname, pw):
            u     = db.get_user(uname)
            token = db.create_session(uname, u["role"])
            db.log_action(uname,"LOGIN",detail=f"Role: {u['role']}")
            resp  = make_response(jsonify(ok=True, role=u["role"], username=uname))
            resp.set_cookie("dcr_token", token, httponly=True,
                            samesite="Lax", secure=IS_RENDER,
                            max_age=db.SESSION_TTL)
            return resp
        db.log_action(uname,"LOGIN_FAIL",detail="Invalid credentials")
        return jsonify(ok=False, error="Invalid username or password"), 401
    return render_login()

@app.route("/logout", methods=["POST"])
def logout():
    token = request.cookies.get("dcr_token")
    if token: db.delete_session(token)
    resp = make_response(redirect("/login"))
    resp.delete_cookie("dcr_token")
    return resp

@app.route("/")
def home():
    return render_dashboard(current_user())

@app.route("/app")
def register():
    pid = request.args.get("p","")
    if not pid: return redirect("/")
    u = current_user()
    proj = db.get_project(pid)
    if not proj: return "Project not found", 404
    return render_register(u, proj)

@app.route("/api/settings/holidays", methods=["GET"])
def api_get_holidays():
    from utils import DEFAULT_HOLIDAYS
    h = db.get_setting("holidays", sorted(DEFAULT_HOLIDAYS))
    return jsonify(holidays=sorted(h) if h else [])

@app.route("/api/settings/holidays", methods=["POST"])
def api_save_holidays():
    u = current_user()
    if not u or u.get("role") not in ("superadmin","admin"):
        return jsonify(error="Superadmin only"), 403
    data = request.get_json(silent=True) or {}
    holidays = data.get("holidays", [])
    if not isinstance(holidays, list): return jsonify(error="Invalid"), 400
    # Validate format
    import datetime
    valid = []
    for h in holidays:
        try: datetime.date.fromisoformat(h); valid.append(h)
        except: pass
    db.save_setting("holidays", sorted(valid))
    from utils import invalidate_holidays_cache
    invalidate_holidays_cache()
    return jsonify(ok=True, count=len(valid))

# ── API: Records ──────────────────────────────────────────────
@app.route("/api/counts/<pid>")
def api_counts(pid):
    dts = db.get_doc_types(pid)
    return jsonify({dt["id"]: db.count_records(pid, dt["id"]) for dt in dts})

@app.route("/api/dashboard_stats")
def api_dashboard_stats():
    stats = db.get_dashboard_stats()
    u = current_user()
    for s in stats:
        s["can_edit"] = can_edit(s["id"])
    return jsonify(stats)

@app.route("/api/next_doc_no/<pid>/<dt_id>")
def api_next_doc_no(pid, dt_id):
    dts    = db.get_doc_types(pid)
    dt     = next((d for d in dts if d["id"] == dt_id), None)
    records = db.get_records(pid, dt_id)
    # Auto-detect prefix from existing records or build from project+dt codes
    if records:
        first_doc = next((r.get("docNo","") for r in records if r.get("docNo","")), "")
        import re as _re
        m = _re.match(r"^(.+?)\s*-\s*\d+", first_doc)
        prefix = m.group(1).strip() if m else (dt["code"] if dt else dt_id)
    else:
        # Build default prefix: PROJECT-DTCODE (e.g. PEM064-DS)
        proj = db.get_project(pid) or {}
        proj_code = proj.get("code", pid).replace("-","")
        dt_code = dt["code"] if dt else dt_id
        prefix = f"{proj_code}-{dt_code}"
    return jsonify(next=get_next_doc_no(prefix, records))

# ── API: Dropdown Lists ───────────────────────────────────────
# ── API: Logos ────────────────────────────────────────────────
# ── API: Users ────────────────────────────────────────────────
@app.route("/api/users")
@require_superadmin
def api_users():
    return jsonify(db.get_all_users())

@app.route("/api/users", methods=["POST"])
@require_superadmin
def api_users_action():
    data   = request.get_json(silent=True) or {}
    action = data.get("action")
    if action == "add":
        uname = data.get("username","").strip().lower()
        pw    = data.get("password","")
        role  = data.get("role","viewer")
        if not uname or not pw:
            return jsonify(ok=False, error="Username and password required"), 400
        db.add_user(uname, pw, role)
        return jsonify(ok=True)
    if action == "delete":
        uname = data.get("username","")
        if uname == "admin":
            return jsonify(ok=False, error="Cannot delete super admin"), 400
        db.delete_user(uname)
        return jsonify(ok=True)
    if action == "change_password":
        db.change_pw(data.get("username",""), data.get("password",""))
        return jsonify(ok=True)
    if action == "assign":
        db.assign_project(data.get("username",""), data.get("project_id",""))
        return jsonify(ok=True)
    if action == "unassign":
        db.unassign_project(data.get("username",""), data.get("project_id",""))
        return jsonify(ok=True)
    return jsonify(ok=False, error="Unknown action"), 400

@app.route("/api/users/<username>/projects")
@require_superadmin
def api_user_projects(username):
    return jsonify(db.get_user_projects(username))

@app.route("/api/whoami")
def api_whoami():
    u = current_user()
    if not u: return jsonify(username="guest", role="guest", projects=[])
    projs = [] if u["role"] == "superadmin" else db.get_user_projects(u["username"])
    return jsonify(username=u["username"], role=u["role"], projects=projs)

@app.route("/api/change_password", methods=["POST"])
def api_change_own_pw():
    u = current_user()
    if not u: return jsonify(error="LOGIN_REQUIRED"), 403
    data = request.get_json(silent=True) or {}
    pw   = data.get("password","")
    if len(pw) < 4: return jsonify(ok=False, error="Min 4 characters"), 400
    db.change_pw(u["username"], pw)
    return jsonify(ok=True)

# ── Export Excel ──────────────────────────────────────────────
@app.route("/api/export_all/<pid>")
def api_export_all(pid):
    u = current_user()
    if u: db.log_action(u["username"],"EXPORT_EXCEL",pid,detail="Export All")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    proj    = db.get_project(pid) or {}
    dts     = db.get_doc_types(pid)
    wb      = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    PRIMARY="1A3A5C"; PL="2563A8"; WHITE="FFFFFF"; ALT="F8FAFC"; OV="FFF5F5"; MUTED="9CA3AF"
    STATUS_XL = {
        "A - Approved":              ("BBF7D0","166534"),
        "B - Approved As Noted":     ("DCFCE7","14532D"),
        "B,C - Approved & Resubmit": ("FED7AA","7C2D12"),
        "C - Revise & Resubmit":     ("FCE7F3","831843"),
        "D - Review not Required":   ("FECACA","7F1D1D"),
        "Under Review":              ("FEF9C3","713F12"),
        "Cancelled":                 ("EF4444","FFFFFF"),
        "Open":                      ("FED7AA","7C2D12"),
        "Closed":                    ("BFDBFE","1E3A5F"),
        "Replied":                   ("D1FAE5","064E3B"),
        "Pending":                   ("E0E7FF","312E81"),
    }
    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(): s=Side(style="thin",color="DDE3ED"); return Border(left=s,right=s,top=s,bottom=s)

    for dt in dts:
        cols    = [c for c in db.get_columns(pid, dt["id"]) if c["visible"]]
        records = db.get_records(pid, dt["id"])
        # always include tab (even if empty)

        ws = wb.create_sheet(title=dt["id"][:31])
        ws.sheet_view.showGridLines = False
        all_cols = [{"col_key":"_sr","label":"Sr."}] + [{"col_key":c["col_key"],"label":c["label"]} for c in cols]
        nc = len(all_cols)

        from openpyxl.utils import get_column_letter as gcl
        def mcell(row, val, bg, fg="FFFFFF", bold=False, sz=11):
            c = ws.cell(row=row, column=1, value=val)
            ws.merge_cells(f"A{row}:{gcl(nc)}{row}")
            c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            return c

        mcell(1, f"{dt['name'].upper()} — {proj.get('name','')} ({proj.get('code','')})",
              PRIMARY, bold=True, sz=12)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 4

        COL_W = {"_sr":5,"docNo":22,"discipline":14,"trade":14,"title":38,"floor":12,
                 "itemRef":16,"issuedDate":13,"expectedReplyDate":14,"actualReplyDate":13,
                 "status":24,"duration":10,"remarks":28,"fileLocation":18}
        CENTER = {"_sr","duration","issuedDate","expectedReplyDate","actualReplyDate"}

        for ci, col in enumerate(all_cols, 1):
            ws.column_dimensions[gcl(ci)].width = COL_W.get(col["col_key"],13)
            c = ws.cell(row=3, column=ci, value=col["label"])
            c.font = Font(bold=True,color=WHITE,size=10,name="Arial")
            c.fill = fill(PRIMARY)
            c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            c.border = thin()
        ws.row_dimensions[3].height = 22

        sr = 1
        if not records:
            # Empty tab — write a "no records" row
            c = ws.cell(row=4, column=1, value="No records in this register")
            c.font = Font(italic=True, size=10, name="Arial", color="9CA3AF")
            ws.merge_cells(f"A4:{gcl(nc)}4")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[4].height = 24
        has_exp_col = any(c["col_key"]=="expectedReplyDate" for c in all_cols)
        for ri, row in enumerate(records):
            rn = 4 + ri
            is_rev = extract_rev(row.get("docNo","")) > 0
            ov     = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"), has_exp_col)
            bg     = OV if ov else (ALT if sr%2==0 else WHITE)
            ws.row_dimensions[rn].height = 18
            for ci, col in enumerate(all_cols, 1):
                key = col["col_key"]
                if key=="_sr":                   val = "" if is_rev else str(sr)
                elif key=="expectedReplyDate":   val = format_date(compute_expected_reply(row.get("issuedDate"),row.get("docNo")))
                elif key=="duration":
                    dur_val = compute_duration(row.get("issuedDate"),row.get("actualReplyDate"))
                    val = str(dur_val) if dur_val is not None else ""
                elif key in ("issuedDate","actualReplyDate"): val = format_date(row.get(key,""))
                else:                            val = str(row.get(key,"") or "")
                if key == "fileLocation" and val and val.startswith("http"):
                    c = ws.cell(row=rn, column=ci)
                    c.value = "View"
                    c.hyperlink = val
                    c.font = Font(size=9,name="Arial",color="2563A8",underline="single")
                else:
                    c = ws.cell(row=rn, column=ci, value=val)
                    if key == "duration" and val == "0":
                        c.value = 0
                c.border = thin()
                c.alignment = Alignment(vertical="center", wrap_text=True,
                                        horizontal="center" if key in CENTER else "left")
                if key=="status" and val:
                    bg2, fg2 = STATUS_XL.get(val, ("F3F4F6","374151"))
                    c.fill = fill(bg2); c.font = Font(bold=True,size=9,name="Arial",color=fg2)
                elif key != "fileLocation" or not (val and val.startswith("http")):
                    c.fill = fill(bg)
                    c.font = Font(size=10,name="Arial",
                                  color=MUTED if is_rev else ("991B1B" if ov else "1E2A3A"))
            if not is_rev: sr += 1

    if not wb.sheetnames:
        ws = wb.create_sheet("Empty"); ws.cell(1,1,"No data")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"{proj.get('code','DCR')}_All_Registers.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/<pid>/<dt_id>")
def api_export(pid, dt_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    proj    = db.get_project(pid) or {}
    cols    = [c for c in db.get_columns(pid, dt_id) if c["visible"]]
    records = db.get_records(pid, dt_id)
    dts     = db.get_doc_types(pid)
    dt      = next((d for d in dts if d["id"] == dt_id), None)

    PRIMARY = "1A3A5C"; PL = "2563A8"; WHITE = "FFFFFF"
    ALT = "F8FAFC"; OV = "FFF5F5"; MUTED = "9CA3AF"

    STATUS_XL = {
        "A - Approved":              ("BBF7D0","166534"),
        "B - Approved As Noted":     ("DCFCE7","14532D"),
        "B,C - Approved & Resubmit": ("FED7AA","7C2D12"),
        "C - Revise & Resubmit":     ("FCE7F3","831843"),
        "D - Review not Required":   ("FECACA","7F1D1D"),
        "Under Review":              ("FEF9C3","713F12"),
        "Cancelled":                 ("EF4444","FFFFFF"),
        "Open":                      ("FED7AA","7C2D12"),
        "Closed":                    ("BFDBFE","1E3A5F"),
        "Replied":                   ("D1FAE5","064E3B"),
        "Pending":                   ("E0E7FF","312E81"),
    }

    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(): s=Side(style="thin",color="DDE3ED"); return Border(left=s,right=s,top=s,bottom=s)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = dt_id; ws.sheet_view.showGridLines = False

    all_cols = [{"col_key":"_sr","label":"Sr."}] + [{"col_key":c["col_key"],"label":c["label"]} for c in cols]
    nc = len(all_cols)

    def mcell(row, val, bg, fg="FFFFFF", bold=False, sz=11, halign="center"):
        c = ws.cell(row=row, column=1, value=val)
        ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
        c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        return c

    mcell(1, f"DOCUMENT CONTROL REGISTER  —  {(dt['name'] if dt else dt_id).upper()}",
          PRIMARY, bold=True, sz=13)
    ws.row_dimensions[1].height = 36

    info = "   |   ".join(f"{k}: {v}" for k,v in [
        ("Project",proj.get("name","")),("Code",proj.get("code","")),
        ("Client",proj.get("client","")),("Consultant",proj.get("mainConsultant","")),
        ("Exported",datetime.datetime.now().strftime("%d/%b/%Y %H:%M"))] if v)
    mcell(2, info, PL, sz=9); ws.row_dimensions[2].height=18
    ws.row_dimensions[3].height = 4

    COL_W = {"_sr":5,"docNo":22,"discipline":14,"trade":14,"title":38,"floor":12,
             "itemRef":16,"issuedDate":13,"expectedReplyDate":14,"actualReplyDate":13,
             "status":24,"duration":10,"remarks":28,"fileLocation":18}
    CENTER = {"_sr","duration","issuedDate","expectedReplyDate","actualReplyDate"}

    for ci, col in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col["col_key"],13)
        c = ws.cell(row=4, column=ci, value=col["label"])
        c.font = Font(bold=True,color=WHITE,size=10,name="Arial")
        c.fill = fill(PRIMARY)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = thin()
    ws.row_dimensions[4].height = 24

    has_exp_col_s = any(c["col_key"]=="expectedReplyDate" for c in cols)
    sr = 1
    for ri, row in enumerate(records):
        rn   = 5 + ri
        is_rev = extract_rev(row.get("docNo","")) > 0
        ov     = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"), has_exp_col_s)
        bg     = OV if ov else (ALT if sr%2==0 else WHITE)
        ws.row_dimensions[rn].height = 20

        for ci, col in enumerate(all_cols, 1):
            key = col["col_key"]
            if key=="_sr":                  val = "" if is_rev else str(sr)
            elif key=="expectedReplyDate":  val = format_date(compute_expected_reply(row.get("issuedDate"),row.get("docNo")))
            elif key=="duration":
                dur_val = compute_duration(row.get("issuedDate"),row.get("actualReplyDate"))
                val = str(dur_val) if dur_val is not None else ""
            elif key=="issuedDate":         val = format_date(row.get(key,""))
            elif key=="actualReplyDate":    val = format_date(row.get(key,""))
            else:                           val = str(row.get(key,"") or "")

            # fileLocation → hyperlink "View"
            if key == "fileLocation" and val and val.startswith("http"):
                c = ws.cell(row=rn, column=ci)
                c.value = "View"
                c.hyperlink = val
                c.font = Font(size=9,name="Arial",color="2563A8",underline="single")
            else:
                c = ws.cell(row=rn, column=ci, value=val)
                # Duration 0 should show as number not empty
                if key == "duration" and val == "0":
                    c.value = 0
                if key=="status" and val:
                    bg2, fg2 = STATUS_XL.get(val, ("F3F4F6","374151"))
                    c.fill = fill(bg2); c.font = Font(bold=True,size=9,name="Arial",color=fg2)
                elif key=="docNo":
                    c.fill = fill(bg)
                    c.font = Font(size=10,name="Consolas",bold=not is_rev,
                                  color=MUTED if is_rev else PRIMARY)
                else:
                    c.fill = fill(bg)
                    c.font = Font(size=10,name="Arial",
                                  color=MUTED if is_rev else ("991B1B" if ov else "1E2A3A"))
            c.border    = thin()
            c.alignment = Alignment(vertical="center", wrap_text=True,
                                    horizontal="center" if key in CENTER else "left")
        if not is_rev: sr += 1

    tot = 5 + len(records)
    real = sum(1 for r in records if extract_rev(r.get("docNo",""))==0)
    mcell(tot, f"TOTAL: {real} documents  |  {len(records)} submissions", PRIMARY, sz=10, halign="left")
    ws.row_dimensions[tot].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"{proj.get('code','DCR')}_{dt_id}_Register.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── PDF Export ────────────────────────────────────────────────

def _build_pdf_for_dt(pid, dt_id, proj, buf=None):
    """Build a PDF for one document type. Returns BytesIO."""
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors as rl_colors

    buf = buf or io.BytesIO()
    cols    = [c for c in db.get_columns(pid, dt_id) if c["visible"]]
    records = db.get_records(pid, dt_id)
    dts     = db.get_doc_types(pid)
    dt      = next((d for d in dts if d["id"] == dt_id), {"name": dt_id, "code": dt_id})

    PAGE = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    DARK   = rl_colors.HexColor("#1A3A5C")
    LIGHT  = rl_colors.HexColor("#F8FAFC")
    OV_COL = rl_colors.HexColor("#FEF2F2")
    REV_COL= rl_colors.HexColor("#F0F9FF")
    ALT    = rl_colors.HexColor("#F0F4F8")

    STATUS_PDF = {
        "A - Approved":              rl_colors.HexColor("#BBF7D0"),
        "B - Approved As Noted":     rl_colors.HexColor("#DCFCE7"),
        "B,C - Approved & Resubmit": rl_colors.HexColor("#FED7AA"),
        "C - Revise & Resubmit":     rl_colors.HexColor("#FCE7F3"),
        "D - Review not Required":   rl_colors.HexColor("#FECACA"),
        "Under Review":              rl_colors.HexColor("#FEF9C3"),
        "Cancelled":                 rl_colors.HexColor("#EF4444"),
        "Open":                      rl_colors.HexColor("#FED7AA"),
        "Closed":                    rl_colors.HexColor("#BFDBFE"),
        "Replied":                   rl_colors.HexColor("#D1FAE5"),
        "Pending":                   rl_colors.HexColor("#E0E7FF"),
    }

    pstyle = ParagraphStyle("cell", fontName="Helvetica", fontSize=7, leading=9,
                             wordWrap="LTR", spaceAfter=0, spaceBefore=0)
    hstyle = ParagraphStyle("hdr", fontName="Helvetica-Bold", fontSize=7.5,
                             leading=10, textColor=rl_colors.white)

    hdr_cols = [{"col_key":"_sr","label":"Sr."}] + [{"col_key":c["col_key"],"label":c["label"]} for c in cols]

    # Column widths (points)
    W_MAP = {"_sr":18,"docNo":68,"discipline":55,"trade":55,"title":110,
             "floor":38,"itemRef":48,"issuedDate":46,"expectedReplyDate":48,
             "actualReplyDate":46,"status":80,"duration":34,"remarks":80,"fileLocation":60}
    page_w = PAGE[0] - 20*mm
    col_ws = [W_MAP.get(c["col_key"], 55) for c in hdr_cols]
    scale  = page_w / sum(col_ws)
    col_ws = [w * scale for w in col_ws]

    # Build header row
    hdr_row = [Paragraph(c["label"], hstyle) for c in hdr_cols]
    data    = [hdr_row]
    row_meta = []   # (bg, is_header)

    sr = 1
    has_exp_col_pdf = any(c["col_key"]=="expectedReplyDate" for c in hdr_cols)
    for row in records:
        is_rev = extract_rev(row.get("docNo","")) > 0
        ov     = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"), has_exp_col_pdf)
        cells  = []
        for c in hdr_cols:
            key = c["col_key"]
            if key == "_sr":
                val = "" if is_rev else str(sr)
            elif key == "expectedReplyDate":
                val = format_date(compute_expected_reply(row.get("issuedDate"), row.get("docNo")))
            elif key == "duration":
                val = str(compute_duration(row.get("issuedDate"), row.get("actualReplyDate")) or "")
            elif key in ("issuedDate","actualReplyDate"):
                val = format_date(row.get(key,""))
            else:
                val = str(row.get(key,"") or "")
            cells.append(Paragraph(val, pstyle))
        data.append(cells)
        if ov:     row_meta.append(OV_COL)
        elif is_rev: row_meta.append(REV_COL)
        elif sr % 2 == 0: row_meta.append(ALT)
        else:      row_meta.append(rl_colors.white)
        if not is_rev: sr += 1

    # Build table style
    ts = [
        ("BACKGROUND", (0,0), (-1,0), DARK),
        ("TEXTCOLOR",  (0,0), (-1,0), rl_colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 7),
        ("ROWBACKGROUND", (0,0), (-1,0), DARK),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ("LEFTPADDING",(0,0),(-1,-1), 2),
        ("RIGHTPADDING",(0,0),(-1,-1), 2),
        ("GRID",       (0,0), (-1,-1), 0.3, rl_colors.HexColor("#CBD5E1")),
    ]
    # Row backgrounds
    for i, bg in enumerate(row_meta, start=1):
        ts.append(("BACKGROUND", (0,i), (-1,i), bg))
    # Status column color
    status_idx = next((i for i,c in enumerate(hdr_cols) if c["col_key"]=="status"), None)
    if status_idx is not None:
        for i, row in enumerate(records, start=1):
            sv = row.get("status","")
            bg = STATUS_PDF.get(sv)
            if bg: ts.append(("BACKGROUND", (status_idx,i), (status_idx,i), bg))

    tbl = Table(data, colWidths=col_ws, repeatRows=1)
    tbl.setStyle(TableStyle(ts))

    proj_name = proj.get("name","") if isinstance(proj, dict) else ""
    title_st  = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=12,
                                textColor=DARK, spaceAfter=4)
    sub_st    = ParagraphStyle("s", fontName="Helvetica", fontSize=9,
                                textColor=rl_colors.HexColor("#64748B"), spaceAfter=8)

    story = [
        Paragraph(f"{dt['name'].upper()} — {proj_name}", title_st),
        Paragraph(f"Project: {proj.get('code','')}  |  Exported: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}", sub_st),
        HRFlowable(width="100%", thickness=1, color=DARK, spaceAfter=6),
        tbl,
    ]
    doc.build(story)
    buf.seek(0)
    return buf


@app.route("/api/export_pdf/<pid>/<dt_id>")
def api_export_pdf(pid, dt_id):
    proj = db.get_project(pid) or {}
    buf  = _build_pdf_for_dt(pid, dt_id, proj)
    fname = f"{proj.get('code','DCR')}_{dt_id}_Register.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


@app.route("/api/export_pdf_all/<pid>")
def api_export_pdf_all(pid):
    from reportlab.platypus import SimpleDocTemplate, PageBreak
    proj = db.get_project(pid) or {}
    dts  = db.get_doc_types(pid)

    # Build each DT as separate PDF bytes then merge
    try:
        from pypdf import PdfWriter
        writer = PdfWriter()
        for dt in dts:
            if not db.count_records(pid, dt["id"]): continue
            single_buf = _build_pdf_for_dt(pid, dt["id"], proj)
            from pypdf import PdfReader
            reader = PdfReader(single_buf)
            for page in reader.pages:
                writer.add_page(page)
        out = io.BytesIO()
        writer.write(out); out.seek(0)
    except ImportError:
        # Fallback: just export first DT or give error
        out = _build_pdf_for_dt(pid, dts[0]["id"], proj) if dts else io.BytesIO()

    fname = f"{proj.get('code','DCR')}_All_Registers.pdf"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


@app.route("/api/import/<pid>/<dt_id>", methods=["POST"])
def api_import(pid, dt_id):
    if not can_edit(pid): return jsonify(error="LOGIN_REQUIRED"), 403
    data = request.get_json(silent=True) or {}
    b64  = data.get("file_b64","")
    ext  = data.get("ext","csv")
    if "," in b64: b64 = b64.split(",",1)[1]

    cols    = db.get_columns(pid, dt_id)
    col_map = {c["label"]: c["col_key"] for c in cols}
    imported = 0

    try:
        if ext in ("xlsx","xls"):
            import openpyxl, datetime as _dt
            wb  = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
            ws  = wb.active; header = None
            for row in ws.iter_rows(values_only=True):
                vals = []
                for v in row:
                    if v is None: vals.append("")
                    elif isinstance(v, (_dt.datetime,_dt.date)):
                        vals.append((v.date() if isinstance(v,_dt.datetime) else v).strftime("%Y-%m-%d"))
                    else: vals.append(str(v).strip())
                if not any(vals): continue
                if header is None:
                    if any(v in col_map or v in ("Sr.","Document No.") for v in vals):
                        header = [col_map.get(v.strip(),v.strip()) for v in vals]
                    continue
                row_data = {header[i]:v for i,v in enumerate(vals)
                            if i<len(header) and header[i] and header[i] not in ("Sr.","sr","")}
                if any(row_data.values()):
                    db.save_record(pid, dt_id, str(uuid.uuid4()), row_data); imported+=1
        else:
            import csv, datetime as _dt
            text   = base64.b64decode(b64).decode("utf-8","ignore")
            header = None
            date_cols = {c["col_key"] for c in cols if c.get("col_type") in ("date","auto_date")}
            for line in csv.reader(io.StringIO(text)):
                if not header:
                    if any(c in line for c in ["Document No.","Sr.","docNo"]):
                        header = [col_map.get(h.strip(),h.strip()) for h in line]
                    continue
                if not any(line): continue
                row_data = {}
                for i,val in enumerate(line):
                    if i<len(header) and header[i] and header[i] not in ("sr","Sr.","Sr",""):
                        v = val.strip()
                        if header[i] in date_cols and v:
                            for fmt in ("%d/%b/%Y","%d-%b-%Y","%Y-%m-%d","%d/%m/%Y"):
                                try: v = _dt.datetime.strptime(v,fmt).strftime("%Y-%m-%d"); break
                                except: pass
                        row_data[header[i]] = v
                if any(row_data.values()):
                    db.save_record(pid, dt_id, str(uuid.uuid4()), row_data); imported+=1
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500

    return jsonify(ok=True, imported=imported)

# ── Phase 2 — Analytics APIs ─────────────────────────────────
# ── Phase 3 — Overdue Digest API ─────────────────────────────
# ── Audit Log API ─────────────────────────────────────────────
@app.route("/api/audit")
def api_audit():
    u = current_user()
    if not u or u.get("role") not in ("superadmin","admin"):
        return jsonify(error="Admin only"), 403
    pid      = request.args.get("pid")
    username = request.args.get("username")
    action   = request.args.get("action")
    offset   = int(request.args.get("offset","0"))
    rows     = db.get_audit_log(project_id=pid, username=username,
                                action=action, limit=100, offset=offset)
    users    = [r["username"] for r in db.get_all_users()]
    actions  = db.get_audit_actions()
    projects = db.get_projects()
    return jsonify(
        rows=[{**dict(r), "ts": str(r["ts"])} for r in rows],
        users=users, actions=actions,
        projects=[{"id":p["id"],"name":p["name"],"code":p["code"]} for p in projects]
    )


if __name__ == "__main__":
    db.init()
    db.cleanup_sessions()
    port = int(os.environ.get("PORT", 5000))
    print(f"[DCR Flask] Running on http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=not IS_RENDER)
