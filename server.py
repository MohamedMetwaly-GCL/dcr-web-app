"""
server.py - DCR v6
Multi-Project | Sessions in DB | Role-based access
"""
import json, os, sys, uuid, csv, io, re, base64, datetime, socket
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse, parse_qs, unquote

sys.path.insert(0, str(Path(__file__).parent))
import modules.database as db
from modules.utils import (compute_expected_reply, compute_duration, is_overdue,
                            format_date, get_next_doc_no, extract_rev, extract_seq,
                            build_doc_no, STATUS_COLORS)

PORT      = int(os.environ.get("PORT", 5000))
IS_RENDER = bool(os.environ.get("RENDER", ""))
SESSION_TTL = 8 * 3600

# ── Auth helpers ──────────────────────────────────────────────
def _get_session(headers):
    cookie = headers.get("Cookie", "")
    for part in cookie.split(";"):
        part = part.strip()
        if part.startswith("dcr_token="):
            token = part[len("dcr_token="):]
            return db.get_session(token)
    return None

def _set_cookie_header(token):
    secure = "; Secure" if IS_RENDER else ""
    return f"dcr_token={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_TTL}{secure}"

def _can_edit(session, project_id):
    """Check if session user can edit given project."""
    if not session: return False
    role = session.get("role","")
    if role == "superadmin": return True
    if role in ("admin","editor"):
        assigned = db.get_user_projects(session["username"])
        return project_id in assigned
    return False

def _can_view(session, project_id):
    """Everyone can view all projects."""
    return True

# ── Handler ───────────────────────────────────────────────────
class DCRHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        origin = self.headers.get("Origin", "")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(body))
        if origin:
            self.send_header("Access-Control-Allow-Origin", origin)
            self.send_header("Access-Control-Allow-Credentials", "true")
        else:
            self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, html, status=200):
        body = html.encode("utf-8") if isinstance(html, str) else html
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def send_redirect(self, location):
        self.send_response(302)
        self.send_header("Location", location)
        self.end_headers()

    def read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length) if length else b""

    def do_OPTIONS(self):
        origin = self.headers.get("Origin", "*")
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", origin)
        self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type,Cookie")
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path   = parsed.path.rstrip("/")
        qs     = parse_qs(parsed.query)
        session = _get_session(self.headers)
        # guest session for unauthenticated users
        if not session:
            session = {"username": "guest", "role": "guest"}

        # ── Static ──
        if path.startswith("/static/"):
            self._send_static(path[8:]); return

        # ── Health / Ping ──
        if path == "/ping":
            self.send_json({"status": "ok", "time": datetime.datetime.now().isoformat()}); return
        if path == "/health":
            try:
                db.get_all_projects()
                self.send_json({"status": "ok", "db": "connected"})
            except Exception as e:
                self.send_json({"status": "error", "db": str(e)}, 503)
            return

        # ── Login page ──
        if path == "/login":
            self.send_html(build_login_page()); return

        # ── Home → Global Dashboard ──
        if path in ("", "/", "/dashboard"):
            self.send_html(build_global_dashboard(session)); return

        # ── Project Register ──
        if path == "/app":
            proj_id = qs.get("p", [""])[0]
            if not proj_id:
                self.send_redirect("/"); return
            self.send_html(build_register_page(session, proj_id)); return

        # ── API ──
        if path.startswith("/api/"):
            self._handle_api_get(path, qs, session); return

        self.send_response(404); self.end_headers()

    def _send_static(self, rel_path):
        static_dir = Path(__file__).parent / "static"
        p = static_dir / rel_path
        ext = p.suffix.lower()
        mime = {".css":"text/css",".js":"application/javascript",
                ".png":"image/png",".jpg":"image/jpeg",
                ".ico":"image/x-icon",".svg":"image/svg+xml"}.get(ext,"application/octet-stream")
        try:
            body = p.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Length", len(body))
            self.send_header("Cache-Control", "max-age=3600")
            self.end_headers()
            self.wfile.write(body)
        except FileNotFoundError:
            self.send_response(404); self.end_headers()

    def _handle_api_get(self, path, qs, session):
        pid = qs.get("p", [""])[0]  # project_id param

        if path == "/api/projects":
            projects = db.get_all_projects()
            result = []
            for p in projects:
                try: data = json.loads(p.get("data") or "{}")
                except: data = {}
                result.append({
                    "id": p["id"], "name": p["name"], "code": p["code"],
                    "client": data.get("client",""),
                    "can_edit": _can_edit(session, p["id"])
                })
            self.send_json(result); return

        if path == "/api/project" and pid:
            self.send_json(db.get_project(pid)); return

        if path == "/api/doc_types" and pid:
            self.send_json(db.get_doc_types(pid)); return

        if path.startswith("/api/records/") and pid:
            dt_id  = path.split("/")[-1]
            search = qs.get("search", [""])[0]
            records = db.get_records(pid, dt_id, search=search)
            cols    = db.get_columns(pid, dt_id)
            for row in records:
                row["_expectedReplyDate"] = format_date(compute_expected_reply(row.get("issuedDate"), row.get("docNo")))
                row["_duration"]   = str(compute_duration(row.get("issuedDate"), row.get("actualReplyDate")) or "")
                row["_overdue"]    = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"))
                row["_isRev"]      = extract_rev(row.get("docNo","")) > 0
                row["_issuedDateFmt"]      = format_date(row.get("issuedDate",""))
                row["_actualReplyDateFmt"] = format_date(row.get("actualReplyDate",""))
            self.send_json({"records": records, "columns": cols,
                            "count": db.get_record_count(pid, dt_id)}); return

        if path == "/api/columns" and pid:
            dt_id = qs.get("dt", [""])[0]
            self.send_json(db.get_columns(pid, dt_id)); return

        if path == "/api/dropdown_lists" and pid:
            self.send_json(db.get_all_dropdown_lists(pid)); return

        if path.startswith("/api/next_doc_no/") and pid:
            dt_id = path.split("/")[-1]
            dt = next((d for d in db.get_doc_types(pid) if d["id"] == dt_id), None)
            prefix = dt["code"] if dt else dt_id
            records = db.get_records(pid, dt_id)
            self.send_json({"next": get_next_doc_no(prefix, records)}); return

        if path == "/api/counts" and pid:
            counts = {dt["id"]: db.get_record_count(pid, dt["id"]) for dt in db.get_doc_types(pid)}
            self.send_json(counts); return

        if path.startswith("/api/logo/") and pid:
            key  = unquote(path.split("/")[-1])
            data = db.get_logo(pid, key)
            if data:
                raw = base64.b64decode(data)
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.send_header("Content-Length", len(raw))
                self.end_headers(); self.wfile.write(raw)
            else:
                self.send_response(404); self.end_headers()
            return

        if path.startswith("/api/export/") and pid:
            dt_id = path.split("/")[-1]
            self._handle_export(pid, dt_id); return

        # ── Super admin APIs ──
        if path == "/api/users/list":
            if session.get("role") != "superadmin":
                self.send_json({"error": "Forbidden"}, 403); return
            self.send_json(db.get_all_users()); return

        if path == "/api/users/projects" and pid:
            if session.get("role") != "superadmin":
                self.send_json({"error": "Forbidden"}, 403); return
            assigned = [r["username"] for r in db.get_project_users(pid)]
            self.send_json(assigned); return

        if path == "/api/whoami":
            self.send_json({"username": session["username"], "role": session["role"],
                            "projects": db.get_user_projects(session["username"])
                            if session["role"] not in ("guest","superadmin") else []}); return

        self.send_response(404); self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)
        path   = parsed.path.rstrip("/")
        qs     = parse_qs(parsed.query)
        body   = self.read_body()
        pid    = qs.get("p", [""])[0]

        try: data = json.loads(body) if body else {}
        except: data = {}

        # ── Login ──
        if path == "/api/login":
            uname = data.get("username","").strip().lower()
            pw    = data.get("password","")
            if db.verify_password(uname, pw):
                u = db.get_user(uname)
                token = db.create_session(uname, u["role"])
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Set-Cookie", _set_cookie_header(token))
                out = json.dumps({"ok": True, "role": u["role"], "username": uname}).encode()
                self.send_header("Content-Length", len(out))
                self.end_headers(); self.wfile.write(out)
            else:
                self.send_json({"ok": False, "error": "Invalid username or password"}, 401)
            return

        # ── Logout ──
        if path == "/api/logout":
            s = _get_session(self.headers)
            if s:
                cookie = self.headers.get("Cookie","")
                for part in cookie.split(";"):
                    part = part.strip()
                    if part.startswith("dcr_token="):
                        db.delete_session(part[len("dcr_token="):])
            self.send_response(302)
            self.send_header("Set-Cookie","dcr_token=; Path=/; Max-Age=0")
            self.send_header("Location","/login"); self.end_headers(); return

        session = _get_session(self.headers)
        if not session:
            session = {"username": "guest", "role": "guest"}

        # ── Super admin: manage users ──
        if path == "/api/users":
            if session["role"] != "superadmin":
                self.send_json({"error":"Forbidden"}, 403); return
            action = data.get("action")
            if action == "add":
                uname = data.get("username","").strip().lower()
                role  = data.get("role","viewer")
                pw    = data.get("password","")
                if not uname or not pw:
                    self.send_json({"ok":False,"error":"Username and password required"},400); return
                db.add_user(uname, pw, role)
                self.send_json({"ok":True}); return
            if action == "delete":
                uname = data.get("username","")
                if uname == "admin":
                    self.send_json({"ok":False,"error":"Cannot delete super admin"},400); return
                db.delete_user(uname)
                self.send_json({"ok":True}); return
            if action == "change_password":
                uname = data.get("username","").strip().lower()
                pw    = data.get("password","")
                if uname and pw:
                    db.change_password(uname, pw)
                    self.send_json({"ok":True}); return
                self.send_json({"ok":False,"error":"Invalid"},400); return
            if action == "assign_project":
                db.assign_user_project(data.get("username",""), data.get("project_id",""))
                self.send_json({"ok":True}); return
            if action == "remove_project":
                db.remove_user_project(data.get("username",""), data.get("project_id",""))
                self.send_json({"ok":True}); return
            self.send_json({"ok":False,"error":"Unknown action"},400); return

        # ── Change own password ──
        if path == "/api/change_password":
            if session["role"] == "guest":
                self.send_json({"error":"LOGIN_REQUIRED"},403); return
            pw = data.get("password","")
            if len(pw) < 4:
                self.send_json({"ok":False,"error":"Min 4 characters"},400); return
            db.change_password(session["username"], pw)
            self.send_json({"ok":True}); return

        # ── Super admin: manage projects ──
        if path == "/api/projects/create":
            if session["role"] != "superadmin":
                self.send_json({"error":"Forbidden"},403); return
            proj_id = data.get("id","").strip().upper()
            name    = data.get("name","").strip()
            code    = data.get("code","").strip().upper()
            if not proj_id or not name or not code:
                self.send_json({"ok":False,"error":"ID, name and code required"},400); return
            db.create_project(proj_id, name, code, creator_username=session["username"])
            self.send_json({"ok":True}); return

        if path.startswith("/api/projects/delete/"):
            if session["role"] != "superadmin":
                self.send_json({"error":"Forbidden"},403); return
            proj_id = path.split("/")[-1]
            db.delete_project(proj_id)
            self.send_json({"ok":True}); return

        # ── Project-level write: need can_edit ──
        if not pid:
            self.send_json({"error":"project_id required"},400); return

        if not _can_edit(session, pid):
            self.send_json({"error":"LOGIN_REQUIRED"},403); return

        if path == "/api/project":
            db.save_project(pid, data)
            self.send_json({"ok":True}); return

        if path.startswith("/api/records/"):
            dt_id  = path.split("/")[-1]
            rec_id = data.pop("_id", None) or str(uuid.uuid4())
            for k in list(data.keys()):
                if k.startswith("_"): del data[k]
            db.save_record(pid, dt_id, rec_id, data)
            self.send_json({"ok":True,"id":rec_id}); return

        if path.startswith("/api/delete_record/"):
            rec_id = path.split("/")[-1]
            db.delete_record(rec_id)
            self.send_json({"ok":True}); return

        if path == "/api/doc_types":
            name = data.get("name","").strip()
            code = data.get("code","").strip().upper()
            if name and code:
                db.add_doc_type(pid, code, name, code)
                self.send_json({"ok":True}); return
            self.send_json({"ok":False,"error":"Name and code required"},400); return

        if path.startswith("/api/delete_doc_type/"):
            dt_id = path.split("/")[-1]
            db.delete_doc_type(pid, dt_id)
            self.send_json({"ok":True}); return

        if path == "/api/dropdown_lists/add":
            db.add_dropdown_item(pid, data.get("list_name",""), data.get("item",""))
            self.send_json({"ok":True}); return

        if path == "/api/dropdown_lists/remove":
            db.remove_dropdown_item(pid, data.get("list_name",""), data.get("item",""))
            self.send_json({"ok":True}); return

        if path == "/api/columns/add":
            db.add_column(pid, data["dt_id"], data["col_key"], data["label"],
                          data["col_type"], data.get("list_name"))
            self.send_json({"ok":True}); return

        if path.startswith("/api/columns/visibility/"):
            col_id = int(path.split("/")[-1])
            db.update_col_visibility(col_id, data.get("visible",True))
            self.send_json({"ok":True}); return

        if path.startswith("/api/columns/delete/"):
            col_id = int(path.split("/")[-1])
            db.delete_column(col_id)
            self.send_json({"ok":True}); return

        if path == "/api/logo":
            db.save_logo(pid, data.get("key",""), data.get("data",""))
            self.send_json({"ok":True}); return

        if path == "/api/import_csv":
            self._handle_import_csv(pid, data.get("dt_id",""), data.get("csv_text","")); return

        if path == "/api/import_xlsx":
            self._handle_import_xlsx(pid, data.get("dt_id",""), data.get("file_b64","")); return

        self.send_response(404); self.end_headers()

    def _handle_export(self, pid, dt_id):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        dt      = next((d for d in db.get_doc_types(pid) if d["id"] == dt_id), None)
        cols    = [c for c in db.get_columns(pid, dt_id) if c["visible"]]
        records = db.get_records(pid, dt_id)
        proj    = db.get_project(pid)

        PRIMARY = "1A3A5C"; PRIMARY_L = "2563A8"; WHITE = "FFFFFF"
        ALT_ROW = "F8FAFC"; OVERDUE = "FFF5F5"; REV_CLR = "9CA3AF"

        STATUS_MAP = {
            "A - Approved":              ("BBF7D0","166534"),
            "B - Approved As Noted":     ("DCFCE7","14532D"),
            "B,C - Approved & Resubmit": ("FED7AA","7C2D12"),
            "C - Revise & Resubmit":     ("FCE7F3","831843"),
            "D - Review not Required":   ("FECACA","7F1D1D"),
            "Under Review":              ("FEF9C3","713F12"),
            "Cancelled":                 ("EF4444","FFFFFF"),
            "Open":                      ("FED7AA","7C2D12"),
            "Closed":                    ("BFDBFE","1E3A5F"),
            "Replied":                   ("D1FAE5","064E3B"),
            "Pending":                   ("E0E7FF","312E81"),
        }

        def fill(c): return PatternFill("solid", fgColor=c)
        def bdr():
            s = Side(style="thin", color="DDE3ED")
            return Border(bottom=s, right=s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = dt_id
        ws.sheet_view.showGridLines = False

        # Title
        all_cols = [{"col_key":"_sr","label":"Sr."}] + [{"col_key":c["col_key"],"label":c["label"]} for c in cols]
        ncols = len(all_cols)
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value     = f"DOCUMENT CONTROL REGISTER  —  {(dt['name'] if dt else dt_id).upper()}"
        c.font      = Font(bold=True, color=WHITE, size=13, name="Arial")
        c.fill      = fill(PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Project info
        info = [f"{k}: {v}" for k,v in [
            ("Project", proj.get("name","")), ("Code", proj.get("code","")),
            ("Client", proj.get("client","")), ("Consultant", proj.get("mainConsultant","")),
            ("Contractor", proj.get("contractor","")),
            ("Exported", datetime.datetime.now().strftime("%d/%b/%Y %H:%M"))
        ] if v]
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        c2 = ws["A2"]
        c2.value     = "   |   ".join(info)
        c2.font      = Font(color=WHITE, size=9, name="Arial")
        c2.fill      = fill(PRIMARY_L)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6

        # Headers
        COL_W = {"_sr":6,"docNo":22,"discipline":14,"trade":14,"title":38,"floor":14,
                 "itemRef":16,"issuedDate":13,"expectedReplyDate":14,"actualReplyDate":13,
                 "status":24,"duration":11,"remarks":28,"fileLocation":18}
        for ci, col in enumerate(all_cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col["col_key"], 14)
            c = ws.cell(row=4, column=ci, value=col["label"])
            c.font      = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill      = fill(PRIMARY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = Border(*(Side(style="thin",color="334D6E"),)*0,
                                 left=Side(style="thin",color="334D6E"),
                                 right=Side(style="thin",color="334D6E"),
                                 top=Side(style="thin",color="334D6E"),
                                 bottom=Side(style="thin",color="334D6E"))
        ws.row_dimensions[4].height = 24

        sr = 1
        CENTER_KEYS = {"_sr","duration","issuedDate","expectedReplyDate","actualReplyDate"}
        for ri, row in enumerate(records):
            rn     = 5 + ri
            is_rev = extract_rev(row.get("docNo","")) > 0
            ov     = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"))
            base_bg = OVERDUE if ov else (ALT_ROW if sr%2==0 else WHITE)
            ws.row_dimensions[rn].height = 20

            for ci, col in enumerate(all_cols, 1):
                key = col["col_key"]
                if key == "_sr":            val = "" if is_rev else str(sr)
                elif key == "expectedReplyDate": val = format_date(compute_expected_reply(row.get("issuedDate"), row.get("docNo")))
                elif key == "duration":     val = str(compute_duration(row.get("issuedDate"), row.get("actualReplyDate")) or "")
                elif key == "issuedDate":   val = format_date(row.get(key,""))
                elif key == "actualReplyDate": val = format_date(row.get(key,""))
                else:                       val = str(row.get(key,"") or "")

                c = ws.cell(row=rn, column=ci, value=val)
                c.border    = bdr()
                c.alignment = Alignment(vertical="center",
                    horizontal="center" if key in CENTER_KEYS else "left")

                if key == "status" and val:
                    bg, fg = STATUS_MAP.get(val, ("F3F4F6","374151"))
                    c.fill = fill(bg); c.font = Font(bold=True,size=9,name="Arial",color=fg)
                elif key == "docNo":
                    c.fill = fill(base_bg)
                    c.font = Font(size=10,name="Consolas",bold=not is_rev,
                                  color=REV_CLR if is_rev else PRIMARY)
                else:
                    c.fill = fill(base_bg)
                    c.font = Font(size=10,name="Arial",
                                  color=REV_CLR if is_rev else ("991B1B" if ov else "1E2A3A"))
            if not is_rev: sr += 1

        # Totals row
        tot = 5 + len(records)
        ws.merge_cells(f"A{tot}:{get_column_letter(ncols)}{tot}")
        real = sum(1 for r in records if extract_rev(r.get("docNo",""))==0)
        tc = ws[f"A{tot}"]
        tc.value     = f"TOTAL: {real} documents  |  {len(records)} submissions  |  Exported: {datetime.datetime.now().strftime('%d/%b/%Y %H:%M')}"
        tc.font      = Font(bold=True,color=WHITE,size=10,name="Arial")
        tc.fill      = fill(PRIMARY)
        tc.alignment = Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[tot].height = 22

        buf = io.BytesIO(); wb.save(buf); body = buf.getvalue()
        fname = f"{proj.get('code','DCR')}_{dt_id}_Register.xlsx"
        self.send_response(200)
        self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition",f'attachment; filename="{fname}"')
        self.send_header("Content-Length",len(body))
        self.end_headers(); self.wfile.write(body)

    def _handle_import_xlsx(self, pid, dt_id, file_b64):
        try:
            import openpyxl
            if "," in file_b64: file_b64 = file_b64.split(",",1)[1]
            wb  = openpyxl.load_workbook(io.BytesIO(base64.b64decode(file_b64)), data_only=True)
            ws  = wb.active
            cols_cfg = db.get_columns(pid, dt_id)
            col_map  = {c["label"]: c["col_key"] for c in cols_cfg}
            imported = 0; header = None
            import datetime as _dt
            for row in ws.iter_rows(values_only=True):
                vals = []
                for v in row:
                    if v is None: vals.append("")
                    elif isinstance(v, (_dt.datetime, _dt.date)):
                        vals.append((v.date() if isinstance(v,_dt.datetime) else v).strftime("%Y-%m-%d"))
                    else: vals.append(str(v).strip())
                if not any(vals): continue
                if header is None:
                    if any(v in col_map or v in ("Sr.","docNo","Document No.","Sr") for v in vals):
                        header = [col_map.get(v.strip(), v.strip()) for v in vals]
                    continue
                row_data = {header[i]: v for i,v in enumerate(vals)
                            if i<len(header) and header[i] and header[i] not in ("Sr.","sr","Sr","")}
                if row_data and any(row_data.values()):
                    db.save_record(pid, dt_id, str(uuid.uuid4()), row_data); imported += 1
            self.send_json({"ok":True,"imported":imported})
        except Exception as e:
            self.send_json({"ok":False,"error":str(e)},500)

    def _handle_import_csv(self, pid, dt_id, csv_text):
        import datetime as _dt
        cols    = db.get_columns(pid, dt_id)
        col_map = {c["label"]: c["col_key"] for c in cols}
        date_cols = {c["col_key"] for c in cols if c.get("col_type") in ("date","auto_date")}
        imported = 0; header = None
        for line in csv.reader(io.StringIO(csv_text)):
            if not header:
                if any(c in line for c in ["Document No.","Sr.","docNo","Sr"]):
                    header = [col_map.get(h.strip(), h.strip()) for h in line]
                continue
            if not any(line): continue
            row_data = {}
            for i,val in enumerate(line):
                if i<len(header) and header[i] and header[i] not in ("sr","Sr.","Sr",""):
                    v = val.strip()
                    if header[i] in date_cols and v:
                        for fmt in ("%d/%b/%Y","%d-%b-%Y","%Y-%m-%d","%d/%m/%Y","%m/%d/%Y"):
                            try: v = _dt.datetime.strptime(v,fmt).strftime("%Y-%m-%d"); break
                            except: pass
                    row_data[header[i]] = v
            if row_data and any(row_data.values()):
                db.save_record(pid, dt_id, str(uuid.uuid4()), row_data); imported += 1
        self.send_json({"ok":True,"imported":imported})



# ═══════════════════════════════════════════════════════════════
# HTML PAGES
# ═══════════════════════════════════════════════════════════════

def _topbar(session, extra_btns=""):
    uname = session.get("username","guest")
    role  = session.get("role","guest")
    rbg   = "rgba(240,165,0,.35)" if role=="superadmin" else "rgba(255,255,255,.18)"
    role_lbl = {"superadmin":"SUPER ADMIN","admin":"ADMIN","editor":"EDITOR",
                "viewer":"VIEWER","guest":"GUEST"}.get(role, role.upper())
    login_btn = ('<a href="/login"><button class="tb-btn" style="background:rgba(240,165,0,.3);'
                 'font-weight:700;border-color:rgba(240,165,0,.7)">🔐 Login</button></a>'
                 if uname == "guest" else
                 '<form action="/api/logout" method="post" style="display:inline">'
                 '<button type="submit" class="tb-btn">⏻ Logout</button></form>')
    admin_btn = ('<button class="tb-btn" onclick="openSuperAdmin()" '
                 'style="background:rgba(240,165,0,.2)">⚙ Admin</button>'
                 if role == "superadmin" else "")
    pw_btn = ('' if uname == "guest" else
              '<button class="tb-btn" onclick="openChangePw()">🔑</button>')
    return f"""<div id="topbar">
  <span style="font-size:20px">📋</span>
  <span style="font-weight:700;font-size:14px;letter-spacing:.3px">Document Control Register</span>
  <div class="spacer"></div>
  {extra_btns}
  {admin_btn}
  {pw_btn}
  <span style="color:rgba(255,255,255,.45);padding:0 4px">|</span>
  <span style="color:rgba(255,255,255,.8);font-size:11px">👤 {uname}
    <span style="background:{rbg};border-radius:3px;padding:1px 7px;font-size:9px;font-weight:700">{role_lbl}</span>
  </span>
  {login_btn}
</div>"""


_BASE_CSS = """
<style>
:root{--primary:#1a3a5c;--primary-lt:#2563a8;--accent:#f0a500;--bg:#f0f4f8;
  --white:#fff;--border:#dde3ed;--text:#1e2a3a;--muted:#6b7a94;
  --success:#16a34a;--danger:#ef4444;--warning:#f59e0b;--radius:6px}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:var(--bg);color:var(--text);font-size:13px}
#topbar{background:var(--primary);color:#fff;height:46px;display:flex;align-items:center;
  padding:0 16px;gap:8px;box-shadow:0 2px 8px rgba(0,0,0,.25);z-index:100;flex-shrink:0}
#topbar .spacer{flex:1}
.tb-btn{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.25);color:#fff;
  padding:5px 12px;border-radius:var(--radius);cursor:pointer;font-size:12px;font-family:inherit;
  text-decoration:none;display:inline-block}
.tb-btn:hover{background:rgba(255,255,255,.25)}
.overlay{position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;
  display:flex;align-items:center;justify-content:center;backdrop-filter:blur(3px)}
.overlay.hidden{display:none}
.modal{background:#fff;border-radius:10px;box-shadow:0 24px 64px rgba(0,0,0,.3);
  width:90%;max-width:600px;max-height:90vh;display:flex;flex-direction:column;
  overflow:hidden;animation:mIn .2s ease}
@keyframes mIn{from{transform:translateY(-16px);opacity:0}to{transform:none;opacity:1}}
.modal-hdr{background:var(--primary);color:#fff;padding:13px 18px;display:flex;
  align-items:center;justify-content:space-between;font-weight:700;font-size:13px;flex-shrink:0}
.modal-close{background:none;border:none;color:#fff;font-size:20px;cursor:pointer;opacity:.7;line-height:1}
.modal-close:hover{opacity:1}
.modal-body{padding:18px;overflow-y:auto;flex:1}
.modal-footer{padding:10px 18px;border-top:1px solid var(--border);display:flex;
  justify-content:flex-end;gap:8px;background:var(--bg);flex-shrink:0}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.form-group{display:flex;flex-direction:column;gap:4px}
.form-group.full{grid-column:1/-1}
.form-group label{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.4px}
.form-group input,.form-group select,.form-group textarea{padding:7px 10px;border:1px solid var(--border);
  border-radius:var(--radius);font-family:inherit;font-size:12px;outline:none}
.form-group input:focus,.form-group select:focus{border-color:var(--primary-lt);
  box-shadow:0 0 0 2px rgba(37,99,168,.1)}
.btn{padding:7px 16px;border-radius:var(--radius);cursor:pointer;font-family:inherit;
  font-size:12px;font-weight:600;border:1px solid transparent}
.btn-primary{background:var(--primary);color:#fff}
.btn-primary:hover{background:var(--primary-lt)}
.btn-secondary{background:var(--bg);color:var(--text);border-color:var(--border)}
.btn-secondary:hover{background:var(--border)}
.btn-danger{background:var(--danger);color:#fff}
.btn-success{background:var(--success);color:#fff}
.btn-sm{padding:4px 10px;font-size:11px}
.section-title{font-size:11px;font-weight:700;color:var(--primary);text-transform:uppercase;
  letter-spacing:.5px;margin:14px 0 6px;padding-bottom:4px;border-bottom:2px solid var(--primary)}
#toast{position:fixed;bottom:32px;right:20px;background:var(--primary);color:#fff;
  padding:10px 18px;border-radius:var(--radius);font-size:12px;z-index:9999;
  box-shadow:0 4px 16px rgba(0,0,0,.2);transform:translateY(80px);opacity:0;
  transition:all .3s;pointer-events:none}
#toast.show{transform:none;opacity:1}
#toast.success{background:#16a34a}
#toast.error{background:var(--danger)}
#toast.warning{background:var(--warning);color:#000}
</style>"""


_SHARED_JS = """
<script>
async function api(url, opts={}) {
  const r = await fetch(url, {credentials:'include', headers:{'Content-Type':'application/json'}, ...opts});
  if(r.status===403){
    const d=await r.json().catch(()=>({}));
    if(d.error==='LOGIN_REQUIRED'){window.location='/login'; return null;}
    throw new Error(d.error||'Forbidden');
  }
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}
function openModal(id){document.getElementById(id).classList.remove('hidden')}
function closeModal(id){document.getElementById(id).classList.add('hidden')}
document.addEventListener('DOMContentLoaded',()=>{
  document.querySelectorAll('.overlay').forEach(o=>o.addEventListener('click',e=>{
    if(e.target===o) o.classList.add('hidden');
  }));
});
function toast(msg, type='info'){
  const t=document.getElementById('toast');
  t.textContent=msg; t.className='show '+(type||'');
  clearTimeout(t._t); t._t=setTimeout(()=>t.className='',3200);
}
async function openChangePw(){
  const pw = prompt('Enter new password (min 4 chars):');
  if(!pw||pw.length<4){return;}
  const r = await api('/api/change_password',{method:'POST',body:JSON.stringify({password:pw})});
  if(r&&r.ok) toast('✔ Password changed!','success');
  else toast((r&&r.error)||'Error','error');
}
</script>"""


# ── LOGIN PAGE ────────────────────────────────────────────────
def build_login_page():
    return f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>DCR — Login</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:linear-gradient(135deg,#0f2640,#1a3a5c 60%,#2563a8);
  min-height:100vh;display:flex;align-items:center;justify-content:center}}
.card{{background:#fff;border-radius:16px;box-shadow:0 24px 80px rgba(0,0,0,.4);width:100%;max-width:400px;overflow:hidden}}
.card-hdr{{background:linear-gradient(135deg,#1a3a5c,#2563a8);padding:32px 32px 24px;text-align:center}}
.card-hdr .ico{{font-size:40px;margin-bottom:8px}}
.card-hdr h1{{color:#fff;font-size:20px;font-weight:700}}
.card-hdr p{{color:rgba(255,255,255,.65);font-size:12px;margin-top:4px}}
.card-body{{padding:28px 32px 32px}}
.field{{margin-bottom:16px}}
.field label{{display:block;font-size:11px;font-weight:700;color:#6b7a94;
  text-transform:uppercase;letter-spacing:.4px;margin-bottom:5px}}
.field input{{width:100%;padding:11px 14px;border:1.5px solid #dde3ed;border-radius:8px;
  font-family:inherit;font-size:13px;outline:none;transition:border-color .2s}}
.field input:focus{{border-color:#2563a8;box-shadow:0 0 0 3px rgba(37,99,168,.12)}}
.err{{background:#fef2f2;border:1px solid #fecaca;color:#dc2626;padding:9px 12px;
  border-radius:6px;font-size:12px;margin-bottom:14px;display:none}}
.btn-login{{width:100%;padding:13px;background:linear-gradient(135deg,#1a3a5c,#2563a8);
  color:#fff;border:none;border-radius:8px;font-family:inherit;font-size:14px;
  font-weight:700;cursor:pointer;transition:all .2s}}
.btn-login:hover{{transform:translateY(-1px);box-shadow:0 4px 16px rgba(26,58,92,.4)}}
.hint{{text-align:center;color:#9ca3af;font-size:11px;margin-top:16px}}
</style></head><body>
<div class="card">
  <div class="card-hdr"><div class="ico">📋</div>
    <h1>Document Control Register</h1><p>Sign in to continue</p></div>
  <div class="card-body">
    <div class="err" id="err"></div>
    <div class="field"><label>Username</label>
      <input id="un" type="text" placeholder="Username" autocomplete="username" autofocus></div>
    <div class="field"><label>Password</label>
      <input id="pw" type="password" placeholder="Password" autocomplete="current-password"></div>
    <button class="btn-login" onclick="doLogin()">Sign In →</button>
    <p class="hint">Contact your administrator for credentials</p>
  </div>
</div>
<script>
document.getElementById('pw').addEventListener('keydown',e=>{{if(e.key==='Enter')doLogin()}});
document.getElementById('un').addEventListener('keydown',e=>{{if(e.key==='Enter')document.getElementById('pw').focus()}});
async function doLogin(){{
  const un=document.getElementById('un').value.trim();
  const pw=document.getElementById('pw').value;
  const err=document.getElementById('err');
  err.style.display='none';
  if(!un||!pw){{err.textContent='Please enter username and password';err.style.display='block';return;}}
  const r=await fetch('/api/login',{{method:'POST',credentials:'include',
    headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{username:un,password:pw}})}});
  const d=await r.json();
  if(d.ok){{window.location='/';}}
  else{{err.textContent=d.error||'Invalid credentials';err.style.display='block';}}
}}
</script></body></html>"""


# ── GLOBAL DASHBOARD ─────────────────────────────────────────
def build_global_dashboard(session):
    uname   = session.get("username","guest")
    role    = session.get("role","guest")
    projects = db.get_all_projects()

    # Build stats per project
    stats = []
    for p in projects:
        try: pdata = json.loads(p.get("data") or "{}")
        except: pdata = {}
        dtypes = db.get_doc_types(p["id"])
        total = approved = pending = overdue = 0
        for dt in dtypes:
            recs = db.get_records(p["id"], dt["id"])
            total    += len([r for r in recs if extract_rev(r.get("docNo",""))==0])
            approved += len([r for r in recs if r.get("status","").startswith("A") or "approved" in str(r.get("status","")).lower()])
            pending  += len([r for r in recs if r.get("status","") in ("Under Review","Pending","")])
            overdue  += len([r for r in recs if is_overdue(r.get("issuedDate"), r.get("docNo"), r.get("actualReplyDate"))])
        pct = round(approved/total*100) if total else 0
        logo_left = db.get_logo(p["id"], "logo_left")
        stats.append({
            "id": p["id"], "name": p["name"], "code": p["code"],
            "client": pdata.get("client",""), "contractor": pdata.get("contractor",""),
            "total": total, "approved": approved, "pending": pending,
            "overdue": overdue, "pct": pct,
            "can_edit": _can_edit(session, p["id"]),
            "has_logo": bool(logo_left)
        })

    stats_json = json.dumps(stats)
    total_all    = sum(s["total"]    for s in stats)
    approved_all = sum(s["approved"] for s in stats)
    pending_all  = sum(s["pending"]  for s in stats)
    overdue_all  = sum(s["overdue"]  for s in stats)
    pct_all = round(approved_all/total_all*100) if total_all else 0

    superadmin_btn = '<button class="tb-btn" onclick="openSuperAdmin()" style="background:rgba(240,165,0,.2)">⚙ Admin Panel</button>' if role=="superadmin" else ""

    return f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>DCR — Dashboard</title>
{_BASE_CSS}
<style>
body{{display:flex;flex-direction:column;min-height:100vh}}
.wrap{{max-width:1400px;margin:0 auto;padding:20px 16px;flex:1}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:28px}}
.kpi{{background:#fff;border-radius:8px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.07);
  border-left:4px solid var(--primary);transition:transform .15s}}
.kpi:hover{{transform:translateY(-2px)}}
.kpi.green{{border-left-color:var(--success)}}
.kpi.orange{{border-left-color:var(--warning)}}
.kpi.red{{border-left-color:var(--danger)}}
.kpi-val{{font-size:32px;font-weight:800;color:var(--primary)}}
.kpi.green .kpi-val{{color:var(--success)}}
.kpi.orange .kpi-val{{color:var(--warning)}}
.kpi.red .kpi-val{{color:var(--danger)}}
.kpi-lbl{{font-size:11px;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:.4px;margin-top:4px}}
.proj-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:16px;margin-bottom:28px}}
.proj-card{{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);
  overflow:hidden;transition:transform .15s,box-shadow .15s;text-decoration:none;color:inherit;display:block}}
.proj-card:hover{{transform:translateY(-3px);box-shadow:0 8px 24px rgba(0,0,0,.12)}}
.proj-card-hdr{{background:var(--primary);padding:14px 18px;display:flex;align-items:center;justify-content:space-between}}
.proj-card-body{{padding:16px 18px}}
.proj-stat{{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}}
.prog-bar{{height:6px;background:#eef1f7;border-radius:99px;overflow:hidden;margin-top:8px}}
.prog-fill{{height:100%;border-radius:99px;background:var(--success)}}
.add-proj-card{{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);
  border:2px dashed var(--border);display:flex;align-items:center;justify-content:center;
  min-height:180px;cursor:pointer;transition:all .15s;color:var(--muted);font-size:13px;
  flex-direction:column;gap:8px;text-decoration:none}}
.add-proj-card:hover{{border-color:var(--primary);color:var(--primary);background:#f7faff}}
.charts-row{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}}
.chart-card{{background:#fff;border-radius:8px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.07)}}
canvas{{max-height:240px}}
@media(max-width:768px){{.charts-row{{grid-template-columns:1fr}}.proj-grid{{grid-template-columns:1fr}}}}

/* Super Admin Modal */
.user-row{{display:flex;align-items:center;gap:8px;padding:7px 10px;background:var(--bg);border-radius:4px;font-size:12px;margin-bottom:4px}}
.badge{{display:inline-block;border-radius:10px;padding:2px 8px;font-size:10px;font-weight:700}}
.badge.admin{{background:#fef3c7;color:#92400e}}
.badge.editor{{background:#e0e7ff;color:#3730a3}}
.badge.viewer{{background:#f0fdf4;color:#166534}}
.badge.superadmin{{background:#fef3c7;color:#92400e}}
.proj-assign{{display:flex;flex-wrap:wrap;gap:4px;margin-top:4px}}
.proj-tag{{background:#eff6ff;color:#1e40af;border-radius:4px;padding:2px 8px;font-size:10px;
  display:flex;align-items:center;gap:4px}}
.proj-tag .rm{{cursor:pointer;color:#ef4444}}
</style>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head><body>
{_topbar(session, extra_btns=superadmin_btn)}

<div class="wrap">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:20px;flex-wrap:wrap;gap:12px">
    <div>
      <h2 style="font-size:22px;font-weight:800;color:var(--primary)">📊 Projects Dashboard</h2>
      <p style="color:var(--muted);font-size:12px;margin-top:4px">All projects — read-only overview. Login to edit.</p>
    </div>
    <div style="background:#fff;border-radius:8px;padding:12px 20px;box-shadow:0 1px 4px rgba(0,0,0,.07);text-align:center">
      <div style="font-size:28px;font-weight:800;color:#16a34a">{pct_all}%</div>
      <div style="font-size:10px;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:.4px">Overall Completion</div>
    </div>
  </div>

  <!-- KPI Cards -->
  <div class="kpi-grid">
    <div class="kpi"><div class="kpi-val">{total_all}</div><div class="kpi-lbl">Total Documents</div></div>
    <div class="kpi green"><div class="kpi-val">{approved_all}</div><div class="kpi-lbl">Approved</div></div>
    <div class="kpi orange"><div class="kpi-val">{pending_all}</div><div class="kpi-lbl">Under Review</div></div>
    <div class="kpi red"><div class="kpi-val">{overdue_all}</div><div class="kpi-lbl">Overdue</div></div>
    <div class="kpi"><div class="kpi-val">{len(projects)}</div><div class="kpi-lbl">Projects</div></div>
  </div>

  <!-- Project Cards -->
  <div style="font-size:13px;font-weight:700;color:var(--primary);text-transform:uppercase;
    letter-spacing:.4px;margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid var(--primary)">
    🗂 All Projects
  </div>
  <div class="proj-grid" id="proj-grid"></div>

  <!-- Charts -->
  <div class="charts-row">
    <div class="chart-card">
      <div style="font-size:11px;font-weight:700;color:var(--primary);text-transform:uppercase;
        letter-spacing:.4px;margin-bottom:12px">Documents by Project</div>
      <canvas id="chartProj"></canvas>
    </div>
    <div class="chart-card">
      <div style="font-size:11px;font-weight:700;color:var(--primary);text-transform:uppercase;
        letter-spacing:.4px;margin-bottom:12px">Overall Status</div>
      <canvas id="chartStatus"></canvas>
    </div>
  </div>
</div>

<div id="toast"></div>

<!-- Super Admin Modal -->
<div class="overlay hidden" id="admin-modal">
  <div class="modal" style="max-width:760px">
    <div class="modal-hdr"><span>⚙ Admin Panel</span>
      <button class="modal-close" onclick="closeModal('admin-modal')">✕</button></div>
    <div class="modal-body" id="admin-body"></div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('admin-modal')">Close</button></div>
  </div>
</div>

<!-- New Project Modal -->
<div class="overlay hidden" id="newproj-modal">
  <div class="modal" style="max-width:440px">
    <div class="modal-hdr"><span>➕ New Project</span>
      <button class="modal-close" onclick="closeModal('newproj-modal')">✕</button></div>
    <div class="modal-body">
      <div class="form-grid">
        <div class="form-group"><label>Project ID / Code</label>
          <input id="np-id" placeholder="e.g. 24CP01" style="text-transform:uppercase"></div>
        <div class="form-group"><label>Short Code</label>
          <input id="np-code" placeholder="e.g. CP01" style="text-transform:uppercase"></div>
        <div class="form-group full"><label>Project Name</label>
          <input id="np-name" placeholder="e.g. New Office Building"></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('newproj-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="createProject()">Create Project</button></div>
  </div>
</div>

{_SHARED_JS}
<script>
const STATS = {stats_json};
const ROLE  = '{role}';

// Render project cards
const grid = document.getElementById('proj-grid');
STATS.forEach(p => {{
  const pct_color = p.pct>=80?'#16a34a':p.pct>=50?'#f59e0b':'#ef4444';
  const card = document.createElement('a');
  card.href  = '/app?p='+p.id;
  card.className = 'proj-card';
  card.innerHTML = `
    <div class="proj-card-hdr">
      <div>
        <div style="color:rgba(255,255,255,.65);font-size:10px;font-weight:700">${{p.code}}</div>
        <div style="color:#fff;font-weight:700;font-size:14px;margin-top:2px">${{p.name}}</div>
      </div>
      ${{p.overdue>0?`<span style="background:#ef4444;color:#fff;border-radius:10px;padding:2px 9px;font-size:10px;font-weight:700">⚠ ${{p.overdue}}</span>`:''}}
    </div>
    <div class="proj-card-body">
      ${{p.client?`<div style="font-size:11px;color:var(--muted);margin-bottom:8px">👤 ${{p.client}}</div>`:''}}
      <div class="proj-stat">
        <span style="font-size:11px;color:var(--muted)">Total</span>
        <span style="font-weight:700">${{p.total}}</span>
      </div>
      <div class="proj-stat">
        <span style="font-size:11px;color:var(--muted)">Approved</span>
        <span style="font-weight:700;color:#16a34a">${{p.approved}}</span>
      </div>
      <div class="proj-stat">
        <span style="font-size:11px;color:var(--muted)">Pending</span>
        <span style="font-weight:700;color:#f59e0b">${{p.pending}}</span>
      </div>
      <div class="prog-bar"><div class="prog-fill" style="width:${{p.pct}}%;background:${{pct_color}}"></div></div>
      <div style="font-size:10px;color:${{pct_color}};font-weight:700;margin-top:4px">${{p.pct}}% complete</div>
      ${{p.can_edit?'<div style="margin-top:8px;font-size:10px;color:#2563a8;font-weight:600">✏ You can edit this project</div>':''}}
    </div>`;
  grid.appendChild(card);
}});

// Add project button (superadmin only)
if(ROLE === 'superadmin') {{
  const add = document.createElement('div');
  add.className = 'add-proj-card';
  add.innerHTML = '<span style="font-size:32px">➕</span><span>New Project</span>';
  add.onclick = () => openModal('newproj-modal');
  grid.appendChild(add);
}}

// Charts
new Chart(document.getElementById('chartProj'), {{
  type:'bar',
  data:{{labels:STATS.map(s=>s.code),
    datasets:[
      {{label:'Approved',data:STATS.map(s=>s.approved),backgroundColor:'#16a34a'}},
      {{label:'Pending',data:STATS.map(s=>s.pending),backgroundColor:'#f59e0b'}},
      {{label:'Overdue',data:STATS.map(s=>s.overdue),backgroundColor:'#ef4444'}},
    ]}},
  options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}},scales:{{x:{{stacked:false}},y:{{beginAtZero:true}}}}}}
}});
new Chart(document.getElementById('chartStatus'), {{
  type:'doughnut',
  data:{{labels:['Approved','Pending','Overdue','Other'],
    datasets:[{{
      data:[{approved_all},{pending_all},{overdue_all},{max(0,total_all-approved_all-pending_all-overdue_all)}],
      backgroundColor:['#16a34a','#f59e0b','#ef4444','#9ca3af'],borderWidth:2,borderColor:'#fff'
    }}]}},
  options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}}}}
}});

// Super Admin Panel
async function openSuperAdmin() {{
  const [users, projects] = await Promise.all([
    api('/api/users/list'), api('/api/projects')
  ]);
  if(!users||!projects) return;
  const body = document.getElementById('admin-body');
  body.innerHTML = '';

  // Users section
  const userTitle = document.createElement('div');
  userTitle.className='section-title'; userTitle.textContent='👥 Users';
  body.appendChild(userTitle);

  users.forEach(u => {{
    const row = document.createElement('div'); row.className='user-row';
    const assignedProjs = projects.filter(p=>p.can_edit||u.role==='superadmin');
    row.innerHTML = `
      <span style="flex:1;font-weight:600">👤 ${{u.username}}</span>
      <span class="badge ${{u.role}}">${{u.role.toUpperCase()}}</span>
      ${{u.username!=='admin'?`
        <button onclick="changeUserPw('${{u.username}}')" class="btn btn-secondary btn-sm">🔑 PW</button>
        <button onclick="delUser('${{u.username}}')" class="btn btn-danger btn-sm">✕</button>`:
        '<span style="font-size:10px;color:var(--muted)">(protected)</span>'}}`;
    body.appendChild(row);

    // Project assignment
    if(u.role !== 'superadmin') {{
      const assignDiv = document.createElement('div');
      assignDiv.style.cssText='padding:0 10px 10px 32px;border-bottom:1px solid var(--border);margin-bottom:4px';
      assignDiv.innerHTML='<div style="font-size:10px;color:var(--muted);margin-bottom:4px">Projects assigned:</div>';
      const tags = document.createElement('div'); tags.className='proj-assign'; tags.id='assign-'+u.username;

      const loadTags = async () => {{
        const assigned = await api('/api/users/projects?p=&username='+u.username)
          .catch(()=>api('/api/whoami').then(()=>[]));
        // Use global projects list
        const uProjs = await fetch('/api/users/projects?p=__unused__&uname='+u.username,
          {{credentials:'include'}}).then(r=>r.ok?r.json():[]).catch(()=>[]);
        tags.innerHTML='';
        projects.forEach(p => {{
          const isAssigned = uProjs.includes(p.id)||false;
          // We'll refetch properly
        }});
      }};

      // Simpler: show all projects with checkboxes
      const projList = document.createElement('div');
      projList.style.cssText='display:flex;flex-wrap:wrap;gap:4px;margin-top:4px';
      projects.forEach(p => {{
        const btn = document.createElement('button');
        btn.className='btn btn-secondary btn-sm';
        btn.style.cssText='font-size:10px;padding:2px 8px';
        btn.dataset.projId = p.id;
        btn.dataset.uname  = u.username;
        btn.textContent    = p.code;
        btn.title          = 'Toggle access to '+p.name;
        btn.onclick = async () => {{
          const assigned = btn.classList.contains('assigned');
          const action   = assigned ? 'remove_project' : 'assign_project';
          await api('/api/users',{{method:'POST',body:JSON.stringify({{action,username:u.username,project_id:p.id}})}});
          btn.classList.toggle('assigned');
          btn.style.background = btn.classList.contains('assigned') ? '#1a3a5c' : '';
          btn.style.color      = btn.classList.contains('assigned') ? '#fff' : '';
          toast(btn.classList.contains('assigned')?`✔ ${{u.username}} → ${{p.name}}`:`Removed ${{u.username}} from ${{p.name}}`,'success');
        }};
        projList.appendChild(btn);
      }});
      // Load current assignments
      fetch('/api/users/projects_for?username='+u.username,{{credentials:'include'}})
        .then(r=>r.ok?r.json():[]).then(assigned=>{{
          projList.querySelectorAll('button[data-proj-id]').forEach(btn=>{{
            if(assigned.includes(btn.dataset.projId)){{
              btn.classList.add('assigned');
              btn.style.background='#1a3a5c'; btn.style.color='#fff';
            }}
          }});
        }}).catch(()=>{{}});
      assignDiv.appendChild(projList);
      body.appendChild(assignDiv);
    }}
  }});

  // Add user
  const addTitle = document.createElement('div');
  addTitle.className='section-title'; addTitle.textContent='➕ Add New User';
  body.appendChild(addTitle);
  const addDiv = document.createElement('div');
  addDiv.innerHTML=`
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:8px;align-items:end">
      <div class="form-group"><label>Username</label><input id="nu-name" placeholder="username"></div>
      <div class="form-group"><label>Role</label>
        <select id="nu-role">
          <option value="editor">Editor (can edit assigned)</option>
          <option value="viewer">Viewer (read-only)</option>
          <option value="superadmin">Super Admin</option>
        </select>
      </div>
      <div class="form-group"><label>Password</label><input id="nu-pw" type="password" placeholder="Password"></div>
      <button class="btn btn-primary btn-sm" style="margin-bottom:1px" onclick="addUser()">Add</button>
    </div>`;
  body.appendChild(addDiv);

  openModal('admin-modal');
}}

async function addUser() {{
  const name=document.getElementById('nu-name').value.trim().toLowerCase();
  const role=document.getElementById('nu-role').value;
  const pw=document.getElementById('nu-pw').value;
  if(!name||!pw){{toast('Username and password required','error');return;}}
  const r=await api('/api/users',{{method:'POST',body:JSON.stringify({{action:'add',username:name,role,password:pw}})}});
  if(r&&r.ok){{toast('✔ User added','success');closeModal('admin-modal');openSuperAdmin();}}
  else toast((r&&r.error)||'Error','error');
}}

async function delUser(uname) {{
  if(!confirm('Delete user: '+uname+'?')) return;
  const r=await api('/api/users',{{method:'POST',body:JSON.stringify({{action:'delete',username:uname}})}});
  if(r&&r.ok){{toast('User deleted','warning');closeModal('admin-modal');openSuperAdmin();}}
  else toast((r&&r.error)||'Error','error');
}}

async function changeUserPw(uname) {{
  const pw=prompt('New password for '+uname+':');
  if(!pw) return;
  const r=await api('/api/users',{{method:'POST',body:JSON.stringify({{action:'change_password',username:uname,password:pw}})}});
  if(r&&r.ok) toast('✔ Password changed','success');
  else toast((r&&r.error)||'Error','error');
}}

async function createProject() {{
  const id=document.getElementById('np-id').value.trim().toUpperCase();
  const code=document.getElementById('np-code').value.trim().toUpperCase();
  const name=document.getElementById('np-name').value.trim();
  if(!id||!name||!code){{toast('All fields required','error');return;}}
  const r=await api('/api/projects/create',{{method:'POST',body:JSON.stringify({{id,name,code}})}});
  if(r&&r.ok){{toast('✔ Project created','success');closeModal('newproj-modal');location.reload();}}
  else toast((r&&r.error)||'Error','error');
}}
</script>
</body></html>"""

# ── ADD missing API endpoint for user project assignments ─────
# Patch the _handle_api_get to add /api/users/projects_for
_orig_api_get = DCRHandler._handle_api_get

def _patched_api_get(self, path, qs, session):
    if path == "/api/users/projects_for":
        if session.get("role") != "superadmin":
            self.send_json({"error": "Forbidden"}, 403); return
        uname = qs.get("username", [""])[0]
        self.send_json(db.get_user_projects(uname)); return
    _orig_api_get(self, path, qs, session)

DCRHandler._handle_api_get = _patched_api_get


# ── REGISTER PAGE ─────────────────────────────────────────────
def build_register_page(session, proj_id):
    proj = db.get_project(proj_id)
    if not proj:
        return "<h2>Project not found</h2>"

    uname    = session.get("username","guest")
    role     = session.get("role","guest")
    can_edit = _can_edit(session, proj_id)
    doc_types = db.get_doc_types(proj_id)
    all_lists = db.get_all_dropdown_lists(proj_id)

    logo_left  = db.get_logo(proj_id, "logo_left")
    logo_right = db.get_logo(proj_id, "logo_right")
    logo_html  = ""
    if logo_left:  logo_html += f'<img src="/api/logo/logo_left?p={proj_id}" style="height:52px;object-fit:contain;margin-right:10px">'
    if logo_right: logo_html += f'<img src="/api/logo/logo_right?p={proj_id}" style="height:52px;object-fit:contain;margin-left:auto;margin-right:10px">'

    tabs_html = "".join(
        f'<button class="tab-btn" data-id="{dt["id"]}" onclick="switchTab(\'{dt["id"]}\')">'
        f'<span class="tab-code">{dt["code"]}</span>'
        f'<span class="tab-count" id="cnt-{dt["id"]}">0</span></button>'
        for dt in doc_types
    )

    proj_fields = [("code","Code"),("name","Project Name"),("client","Client"),
                   ("mainConsultant","Consultant"),("mepConsultant","MEP"),
                   ("contractor","Contractor"),("startDate","Start"),("endDate","End"),
                   ("pmo","PMO"),("landlord","Landlord")]

    proj_bar_html = "".join(
        f'<div class="pf"><span class="pf-lbl">{lbl}</span>'
        f'<span class="pf-val" id="pf-{key}" data-key="{key}">{proj.get(key,"—")}</span></div>'
        for key,lbl in proj_fields
    )

    status_colors_js = json.dumps(STATUS_COLORS)
    edit_toolbar = (f'<button class="tool-btn" onclick="openAddRecord()">➕ Add Document</button>'
                    f'<button class="tool-btn purple" onclick="manageColumns()">⚙ Columns</button>'
                    f'<button class="tool-btn" onclick="openSettings()">📋 Lists</button>'
                    f'<button class="tool-btn" onclick="openProjectModal()">🏗 Project Info</button>'
                    if can_edit else
                    '<span style="font-size:11px;color:rgba(255,255,255,.5);padding:4px 8px">'
                    '👁 Read-only mode — <a href="/login" style="color:#f0a500">Login to edit</a></span>')

    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>DCR — {proj.get('name','Register')}</title>
{_BASE_CSS}
<style>
body{{height:100vh;display:flex;flex-direction:column;overflow:hidden}}
#projbar{{background:#fff;border-bottom:2px solid var(--primary);padding:4px 14px;
  display:flex;align-items:center;gap:0;flex-wrap:nowrap;overflow-x:auto;flex-shrink:0}}
.pf{{display:flex;flex-direction:column;padding:0 10px;border-right:1px solid var(--border)}}
.pf:last-of-type{{border-right:none}}
.pf-lbl{{font-size:9px;font-weight:700;color:var(--primary);text-transform:uppercase;letter-spacing:.4px}}
.pf-val{{font-size:11px;color:var(--text);font-weight:500;white-space:nowrap}}
#tabs-bar{{background:#0f2640;display:flex;align-items:center;overflow-x:auto;
  flex-shrink:0;padding:0 8px;scrollbar-width:thin}}
#tabs-bar::-webkit-scrollbar{{height:3px}}
#tabs-bar::-webkit-scrollbar-thumb{{background:rgba(255,255,255,.3)}}
.tab-btn{{display:flex;align-items:center;gap:6px;padding:9px 13px;background:transparent;
  border:none;border-bottom:2px solid transparent;color:rgba(255,255,255,.55);
  cursor:pointer;font-family:inherit;font-size:11px;font-weight:600;white-space:nowrap;transition:all .15s}}
.tab-btn:hover{{color:#fff;background:rgba(255,255,255,.08)}}
.tab-btn.active{{color:#fff;border-bottom-color:var(--accent)}}
.tab-count{{background:rgba(255,255,255,.2);border-radius:10px;padding:1px 7px;font-size:10px;font-weight:700}}
.tab-btn.active .tab-count{{background:var(--accent);color:#000}}
.tab-add{{padding:6px 10px;background:rgba(255,255,255,.1);border:1px dashed rgba(255,255,255,.35);
  color:rgba(255,255,255,.7);border-radius:4px;cursor:pointer;font-size:16px;margin-left:6px;flex-shrink:0}}
.tab-add:hover{{background:rgba(255,255,255,.2)}}
#toolbar{{background:#fff;border-bottom:1px solid var(--border);padding:6px 12px;
  display:flex;align-items:center;gap:6px;flex-shrink:0;flex-wrap:wrap}}
.tool-btn{{display:flex;align-items:center;gap:4px;padding:5px 11px;background:var(--bg);
  border:1px solid var(--border);border-radius:var(--radius);cursor:pointer;font-size:11px;
  font-family:inherit;color:var(--text);transition:all .15s;white-space:nowrap}}
.tool-btn:hover{{background:var(--primary);color:#fff;border-color:var(--primary)}}
.tool-btn.green:hover{{background:#059669;border-color:#059669}}
.tool-btn.purple:hover{{background:#7c3aed;border-color:#7c3aed}}
.tool-btn.teal:hover{{background:#0891b2;border-color:#0891b2}}
#search-box{{flex:1;min-width:160px;max-width:280px;padding:5px 10px 5px 28px;
  border:1px solid var(--border);border-radius:var(--radius);font-family:inherit;font-size:12px;
  outline:none;background:#fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='%236b7a94' stroke-width='2'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cpath d='m21 21-4.35-4.35'/%3E%3C/svg%3E") no-repeat 7px center}}
#search-box:focus{{border-color:var(--primary-lt);box-shadow:0 0 0 2px rgba(37,99,168,.12)}}
#main{{flex:1;overflow:hidden;display:flex;flex-direction:column}}
#table-wrap{{flex:1;overflow:auto}}
table{{width:100%;border-collapse:collapse;min-width:900px;font-size:12px}}
thead{{position:sticky;top:0;z-index:10}}
th{{background:var(--primary);color:#fff;padding:8px;text-align:left;font-weight:600;
  white-space:nowrap;border-right:1px solid rgba(255,255,255,.12);cursor:pointer;
  user-select:none;position:relative}}
th:hover{{background:var(--primary-lt)}}
.filter-row th{{background:#eef1f7;padding:2px 4px;cursor:default;position:sticky;top:33px;z-index:9}}
.filter-row th:hover{{background:#eef1f7}}
.filter-row input,.filter-row select{{width:100%;padding:3px 6px;border:1px solid var(--border);
  border-radius:3px;font-size:10px;font-family:inherit;background:#fff;outline:none}}
td{{padding:5px 8px;border-bottom:1px solid #edf0f5;border-right:1px solid #f3f4f6;
  vertical-align:middle;max-width:220px;word-break:break-word}}
tr:hover td{{background:rgba(37,99,168,.04)}}
tr.overdue td{{background:#fff5f5}}
tr.rev-row td{{color:var(--muted)}}
tr.alt td{{background:#fafbfd}}
.sr-cell{{text-align:center;color:var(--muted);font-size:10px;width:36px}}
.chk-cell{{text-align:center;width:30px;padding:4px!important}}
.chk-cell input{{width:14px;height:14px;cursor:pointer;accent-color:var(--primary)}}
.actions-cell{{white-space:nowrap;width:68px}}
.act-btn{{padding:2px 7px;border:1px solid var(--border);background:#fff;border-radius:3px;cursor:pointer;font-size:11px}}
.act-btn:hover{{background:var(--primary);color:#fff;border-color:var(--primary)}}
.act-btn.del:hover{{background:var(--danger);border-color:var(--danger)}}
.status-badge{{display:inline-block;border-radius:10px;padding:2px 9px;font-size:10px;font-weight:700}}
.file-link{{color:var(--primary-lt);text-decoration:underline;cursor:pointer;font-size:11px}}
.overdue-date{{color:#dc2626;font-weight:700}}
#bulk-bar{{display:none;background:#1a3a5c;color:#fff;padding:6px 14px;
  align-items:center;gap:10px;font-size:12px;flex-shrink:0}}
#bulk-bar.show{{display:flex}}
#statusbar{{background:var(--primary);color:rgba(255,255,255,.75);padding:3px 14px;
  font-size:10px;display:flex;gap:18px;flex-shrink:0}}
.col-resizer{{position:absolute;right:0;top:0;bottom:0;width:6px;cursor:col-resize;z-index:1}}
.col-resizer:hover,.col-resizer.resizing{{background:var(--accent)}}
.ms-container{{border:1px solid var(--border);border-radius:var(--radius);min-height:36px;
  padding:4px;background:#fff;cursor:pointer;position:relative}}
.ms-tag{{display:inline-flex;align-items:center;gap:3px;background:var(--primary);
  color:#fff;border-radius:3px;padding:2px 7px;font-size:10px;margin:2px}}
.ms-tag .rm{{cursor:pointer;opacity:.7}}
.ms-placeholder{{color:var(--muted);font-size:11px;padding:2px 4px}}
.ms-dropdown{{position:absolute;left:0;right:0;top:100%;background:#fff;
  border:1px solid var(--border);border-radius:var(--radius);z-index:200;
  max-height:200px;overflow-y:auto;box-shadow:0 4px 16px rgba(0,0,0,.12);margin-top:2px}}
.ms-option{{padding:7px 10px;cursor:pointer;font-size:11px;display:flex;align-items:center;gap:7px}}
.ms-option:hover{{background:var(--bg)}}
.ms-option.sel{{background:#eff6ff}}
.empty{{text-align:center;padding:60px 20px;color:var(--muted)}}
.seq-warn{{background:#fef3c7;border:1px solid #f59e0b;border-radius:var(--radius);
  padding:7px 10px;font-size:11px;color:#92400e;display:none;margin-bottom:8px}}
.settings-list{{list-style:none;display:flex;flex-direction:column;gap:4px;
  max-height:200px;overflow-y:auto;border:1px solid var(--border);border-radius:var(--radius);padding:4px}}
.settings-item{{display:flex;align-items:center;gap:8px;padding:4px 8px;
  background:var(--bg);border-radius:3px;font-size:11px}}
.settings-item .name{{flex:1}}
.settings-item button{{padding:2px 8px;font-size:10px;border:1px solid var(--border);
  background:#fff;border-radius:3px;cursor:pointer}}
.settings-item button:hover{{background:var(--danger);color:#fff;border-color:var(--danger)}}
.add-item-row{{display:flex;gap:6px;margin-top:6px}}
.add-item-row input{{flex:1;padding:5px 8px;border:1px solid var(--border);border-radius:3px;
  font-size:11px;font-family:inherit;outline:none}}
@media(max-width:768px){{
  #projbar{{flex-wrap:nowrap;padding:3px 6px}}
  .pf{{padding:0 6px;min-width:72px}}
  #tabs-bar{{padding:0 4px}}
  .tab-btn{{padding:7px 8px;font-size:10px}}
  #toolbar{{padding:4px 6px;gap:4px}}
  .tool-btn{{padding:4px 8px;font-size:10px}}
  table{{font-size:11px}}
  td,th{{padding:4px 5px}}
  .modal{{width:97%;max-height:96vh}}
  .form-grid{{grid-template-columns:1fr}}
}}
</style>
</head><body>

{_topbar(session, extra_btns=f'<a href="/" class="tb-btn">📊 Dashboard</a>')}

<div id="projbar">
  {logo_html}
  {proj_bar_html}
  {'<button style="margin-left:auto;background:var(--primary);color:#fff;border:none;padding:5px 12px;border-radius:var(--radius);cursor:pointer;font-size:11px;font-family:inherit;flex-shrink:0" onclick="openProjectModal()">✏ Edit</button>' if can_edit else ''}
</div>

<div id="tabs-bar">
  {tabs_html}
  {'<button class="tab-add" onclick="openAddDocType()" title="Add Type">＋</button>' if can_edit else ''}
</div>

<div id="toolbar">
  {edit_toolbar}
  <button class="tool-btn green" onclick="doExport()">📥 Export Excel</button>
  {'<button class="tool-btn teal" onclick="openImport()">📤 Import</button>' if can_edit else ''}
  <input type="text" id="search-box" placeholder="Search..." oninput="applySearch()">
</div>

<div id="bulk-bar">
  <span id="bulk-count">0 selected</span>
  <button class="btn btn-danger btn-sm" onclick="bulkDelete()">🗑 Delete Selected</button>
  <button class="btn btn-secondary btn-sm" onclick="clearSelection()" style="background:rgba(255,255,255,.15);color:#fff;border-color:rgba(255,255,255,.3)">✕ Cancel</button>
</div>

<div id="main">
  <div id="table-wrap">
    <table id="reg-table"><thead id="t-head"></thead><tbody id="t-body"></tbody></table>
    <div class="empty hidden" id="empty-state">
      <div style="font-size:48px;margin-bottom:12px">📁</div>
      <h3>No records found</h3>
      <p>{'Add your first document using the button above.' if can_edit else 'No records in this register yet.'}</p>
    </div>
  </div>
</div>

<div id="statusbar">
  <span id="s-total">Total: 0</span>
  <span id="s-showing">Showing: 0</span>
  <span id="s-overdue">Overdue: 0</span>
  <span style="margin-left:auto" id="s-clock"></span>
</div>

<div id="toast"></div>

<!-- ADD/EDIT RECORD -->
<div class="overlay hidden" id="rec-modal">
  <div class="modal" style="max-width:820px">
    <div class="modal-hdr"><span id="rec-modal-title">Add Document</span>
      <button class="modal-close" onclick="closeModal('rec-modal')">✕</button></div>
    <div class="modal-body">
      <div class="seq-warn" id="seq-warn">⚠ Sequence gap detected</div>
      <div class="form-grid" id="rec-form"></div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('rec-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="saveRecord()">Save Document</button></div>
  </div>
</div>

<!-- PROJECT INFO -->
<div class="overlay hidden" id="proj-modal">
  <div class="modal" style="max-width:820px">
    <div class="modal-hdr"><span>🏗 Project Information</span>
      <button class="modal-close" onclick="closeModal('proj-modal')">✕</button></div>
    <div class="modal-body" id="proj-modal-body"></div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('proj-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="saveProject()">Save Project</button></div>
  </div>
</div>

<!-- SETTINGS (dropdown lists) -->
<div class="overlay hidden" id="settings-modal">
  <div class="modal">
    <div class="modal-hdr"><span>📋 Dropdown Lists</span>
      <button class="modal-close" onclick="closeModal('settings-modal')">✕</button></div>
    <div class="modal-body" id="settings-body"></div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('settings-modal')">Close</button></div>
  </div>
</div>

<!-- ADD DOC TYPE -->
<div class="overlay hidden" id="addtype-modal">
  <div class="modal" style="max-width:420px">
    <div class="modal-hdr"><span>Add Document Type</span>
      <button class="modal-close" onclick="closeModal('addtype-modal')">✕</button></div>
    <div class="modal-body">
      <div class="form-grid">
        <div class="form-group"><label>Type Name</label><input id="dt-name" placeholder="e.g. Method Statement"></div>
        <div class="form-group"><label>Code</label><input id="dt-code" placeholder="e.g. MS" style="text-transform:uppercase"></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('addtype-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="saveDocType()">Add</button></div>
  </div>
</div>

<!-- COLUMNS -->
<div class="overlay hidden" id="cols-modal">
  <div class="modal">
    <div class="modal-hdr"><span>⚙ Manage Columns</span>
      <button class="modal-close" onclick="closeModal('cols-modal')">✕</button></div>
    <div class="modal-body" id="cols-body"></div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="openAddColumn()">+ Add Column</button>
      <button class="btn btn-primary" onclick="closeModal('cols-modal');loadRecords()">Done</button></div>
  </div>
</div>

<!-- ADD COLUMN -->
<div class="overlay hidden" id="addcol-modal">
  <div class="modal" style="max-width:460px">
    <div class="modal-hdr"><span>Add Column</span>
      <button class="modal-close" onclick="closeModal('addcol-modal')">✕</button></div>
    <div class="modal-body">
      <div class="form-grid">
        <div class="form-group full"><label>Column Name</label>
          <input id="col-name" placeholder="e.g. Review Duration"></div>
        <div class="form-group full"><label>Type</label>
          <select id="col-type" onchange="onColTypeChange(this.value)">
            <option value="text">📝 Text</option>
            <option value="number">🔢 Number</option>
            <option value="date">📅 Date</option>
            <option value="dropdown">📋 Dropdown List</option>
            <option value="link">🔗 Hyperlink</option>
            <option value="duration_calc">⏱ Duration (working days)</option>
          </select>
        </div>
        <div class="form-group full" id="col-list-grp" style="display:none">
          <label>Dropdown Source</label><select id="col-list-src"></select></div>
        <div class="form-group full" id="col-dur-grp" style="display:none">
          <label>Start Date Column</label><select id="col-dur-start-sel"></select></div>
        <div class="form-group full" id="col-dur-end-grp" style="display:none">
          <label>End Date Column</label><select id="col-dur-end-sel"></select></div>
        <div class="form-group full" id="col-dur-info" style="display:none">
          <div style="background:#f0f4fa;border-radius:6px;padding:9px 12px;font-size:11px;color:#1a3a5c">
            ⏱ Working days between selected dates — excludes Fridays + Egyptian holidays
          </div>
        </div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('addcol-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="saveAddColumn()">Add Column</button></div>
  </div>
</div>

<!-- IMPORT -->
<div class="overlay hidden" id="import-modal">
  <div class="modal" style="max-width:500px">
    <div class="modal-hdr"><span>📤 Import Excel / CSV</span>
      <button class="modal-close" onclick="closeModal('import-modal')">✕</button></div>
    <div class="modal-body">
      <p style="font-size:12px;color:var(--muted);margin-bottom:12px">
        Select an Excel (.xlsx) or CSV file. Column headers must match register columns.</p>
      <input type="file" id="import-file" accept=".csv,.xlsx,.xls" style="font-size:12px">
    </div>
    <div class="modal-footer">
      <button class="btn btn-secondary" onclick="closeModal('import-modal')">Cancel</button>
      <button class="btn btn-primary" onclick="doImport()">Import</button></div>
  </div>
</div>

{_SHARED_JS}
<script>
const PID = '{proj_id}';
const ROLE = '{role}';
const CAN_EDIT = {'true' if can_edit else 'false'};
const STATUS_COLORS = {status_colors_js};

const state = {{
  activeTab: null, docTypes: [], columns: [], allRecords: null,
  sortCol: null, sortDir: 'asc', colFilters: {{}}, editingId: null, allLists: {{}}
}};

const tabParam = new URLSearchParams(window.location.search).get('tab');

function apiP(url, opts={{}}) {{
  const sep = url.includes('?') ? '&' : '?';
  return api(url + sep + 'p=' + PID, opts);
}}

async function init() {{
  await Promise.all([loadDocTypes(), loadLists()]);
  updateClock(); setInterval(updateClock, 60000);
}}

function updateClock() {{
  document.getElementById('s-clock').textContent = new Date().toLocaleString('en-GB');
}}

async function loadDocTypes() {{
  const types = await apiP('/api/doc_types');
  if(!types) return;
  state.docTypes = types;
  renderTabs(types);
  const counts = await apiP('/api/counts');
  if(counts) types.forEach(dt => {{
    const el = document.getElementById('cnt-'+dt.id);
    if(el) el.textContent = counts[dt.id]||0;
  }});
  const target = tabParam || (types.length ? types[0].id : null);
  if(target) switchTab(target);
}}

let _listsPromise = null;
async function loadLists(force=false) {{
  if(!_listsPromise||force) {{
    _listsPromise = apiP('/api/dropdown_lists').then(d=>{{ state.allLists=d||{{}}; }});
  }}
  return _listsPromise;
}}

function renderTabs(types) {{
  const bar = document.getElementById('tabs-bar');
  bar.querySelectorAll('.tab-btn').forEach(b=>b.remove());
  const addBtn = bar.querySelector('.tab-add');
  types.forEach(dt => {{
    const btn = document.createElement('button');
    btn.className = 'tab-btn'+(dt.id===state.activeTab?' active':'');
    btn.dataset.id = dt.id;
    btn.innerHTML = `<span class="tab-code">${{dt.code}}</span><span class="tab-count" id="cnt-${{dt.id}}">0</span>`;
    btn.title = dt.name;
    btn.onclick = ()=>switchTab(dt.id);
    if(CAN_EDIT) btn.oncontextmenu = e=>{{ e.preventDefault(); tabCtxMenu(dt.id,e); }};
    bar.insertBefore(btn, addBtn||null);
  }});
}}

function switchTab(id) {{
  state.activeTab=id; state.colFilters={{}}; state.sortCol=null; state.allRecords=null;
  document.getElementById('search-box').value='';
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.toggle('active',b.dataset.id===id));
  loadRecords();
}}

function tabCtxMenu(id, e) {{
  const old=document.getElementById('tab-ctx'); if(old) old.remove();
  const menu=document.createElement('div');
  menu.id='tab-ctx';
  menu.style.cssText=`position:fixed;left:${{e.clientX}}px;top:${{e.clientY}}px;background:#fff;border:1px solid var(--border);border-radius:6px;box-shadow:0 4px 12px rgba(0,0,0,.15);z-index:9999;min-width:140px;overflow:hidden`;
  menu.innerHTML=`<div onclick="deleteTab('${{id}}')" style="padding:8px 14px;cursor:pointer;font-size:12px;color:#ef4444" onmouseover="this.style.background='#fef2f2'" onmouseout="this.style.background=''">🗑 Delete Type</div>`;
  document.body.appendChild(menu);
  setTimeout(()=>document.addEventListener('click',()=>menu.remove(),{{once:true}}),10);
}}

async function deleteTab(id) {{
  if(!confirm('Delete this document type and ALL its records?')) return;
  await apiP('/api/delete_doc_type/'+id, {{method:'POST',body:'{{}}'}});
  if(state.activeTab===id) state.activeTab=null;
  await loadDocTypes(); toast('Document type deleted','warning');
}}

async function loadRecords() {{
  if(!state.activeTab) return;
  const tbody=document.getElementById('t-body');
  if(tbody && !state.allRecords) tbody.innerHTML='<tr><td colspan="20" style="text-align:center;padding:30px;color:var(--muted)">⏳ Loading...</td></tr>';
  const search=document.getElementById('search-box').value.trim();
  const data=await apiP('/api/records/'+state.activeTab+(search?'&search='+encodeURIComponent(search):''));
  if(!data) return;
  state.allRecords=data.records; state.columns=data.columns.filter(c=>c.visible);
  const cnt=document.getElementById('cnt-'+state.activeTab);
  if(cnt) cnt.textContent=data.count;
  buildTableHead(); requestAnimationFrame(initColResize); renderRows();
}}

function buildTableHead() {{
  const head=document.getElementById('t-head'); head.innerHTML='';
  const hr=document.createElement('tr');
  const chkTh=document.createElement('th'); chkTh.className='chk-cell';
  if(CAN_EDIT) chkTh.innerHTML='<input type="checkbox" id="chk-all" onchange="toggleSelectAll(this.checked)">';
  hr.appendChild(chkTh);
  const srTh=document.createElement('th'); srTh.textContent='Sr.'; srTh.style.width='38px'; srTh.style.cursor='default';
  hr.appendChild(srTh);
  state.columns.forEach(col=>{{
    const th=document.createElement('th'); th.dataset.key=col.col_key;
    th.innerHTML=col.label+(state.sortCol===col.col_key?(state.sortDir==='asc'?' ↑':' ↓'):'');
    if(!['auto_date','auto_num'].includes(col.col_type)) th.onclick=()=>sortBy(col.col_key);
    else th.style.cursor='default';
    hr.appendChild(th);
  }});
  const actTh=document.createElement('th'); actTh.textContent='Actions'; actTh.style.width='68px'; actTh.style.cursor='default';
  hr.appendChild(actTh);
  head.appendChild(hr);
  const fr=document.createElement('tr'); fr.className='filter-row';
  fr.appendChild(document.createElement('th')); fr.appendChild(document.createElement('th'));
  state.columns.forEach(col=>{{
    const th=document.createElement('th');
    if(['auto_date','auto_num'].includes(col.col_type)){{fr.appendChild(th);return;}}
    if(col.col_type==='dropdown'&&col.list_name){{
      const sel=document.createElement('select');
      const opts=state.allLists[col.list_name]||[];
      sel.innerHTML='<option value="">All</option>'+opts.map(o=>`<option ${{state.colFilters[col.col_key]===o?'selected':''}}>${{o}}</option>`).join('');
      sel.onchange=()=>{{state.colFilters[col.col_key]=sel.value;renderRows();}};
      th.appendChild(sel);
    }}else{{
      const inp=document.createElement('input'); inp.value=state.colFilters[col.col_key]||'';
      inp.oninput=()=>{{state.colFilters[col.col_key]=inp.value;renderRows();}};
      th.appendChild(inp);
    }}
    fr.appendChild(th);
  }});
  fr.appendChild(document.createElement('th'));
  head.appendChild(fr);
}}

function renderRows() {{
  const body=document.getElementById('t-body'); body.innerHTML='';
  let rows=state.allRecords.filter(row=>{{
    for(const [k,v] of Object.entries(state.colFilters)){{
      if(v&&!String(row[k]||'').toLowerCase().includes(v.toLowerCase())) return false;
    }}
    return true;
  }});
  if(state.sortCol) rows.sort((a,b)=>{{
    const va=String(a[state.sortCol]||'').toLowerCase();
    const vb=String(b[state.sortCol]||'').toLowerCase();
    return state.sortDir==='asc'?(va>vb?1:-1):(va<vb?1:-1);
  }});
  document.getElementById('empty-state').classList.toggle('hidden',rows.length>0);
  let sr=1;
  rows.forEach((row,idx)=>{{
    const tr=document.createElement('tr');
    if(row._overdue) tr.classList.add('overdue');
    else if(row._isRev) tr.classList.add('rev-row');
    else if(idx%2===1) tr.classList.add('alt');
    const tdChk=document.createElement('td'); tdChk.className='chk-cell';
    if(CAN_EDIT){{const cb=document.createElement('input');cb.type='checkbox';cb.dataset.id=row._id;cb.onchange=updateBulkBar;tdChk.appendChild(cb);}}
    tr.appendChild(tdChk);
    const tdSr=document.createElement('td'); tdSr.className='sr-cell'; tdSr.textContent=row._isRev?'':sr;
    tr.appendChild(tdSr);
    state.columns.forEach(col=>{{
      const td=document.createElement('td'); const key=col.col_key;
      let val='';
      if(key==='expectedReplyDate'){{val=row._expectedReplyDate||'';if(row._overdue&&val)td.classList.add('overdue-date');}}
      else if(key==='duration'){{val=row._duration||'';}}
      else if(col.col_type==='duration_calc'){{
        const parts=(col.list_name||'issuedDate,actualReplyDate').split(',');
        val=calcWorkingDays(row[parts[0].trim()]||'',row[parts[1].trim()]||'');
      }}
      else if(key==='issuedDate'){{val=row._issuedDateFmt||'';}}
      else if(key==='actualReplyDate'){{val=row._actualReplyDateFmt||'';}}
      else if(key==='status'){{
        val=row[key]||'';
        if(val){{
          td.innerHTML=val.split(',').map(s=>{{
            s=s.trim();const [bg,fg]=STATUS_COLORS[s]||['e5e7eb','374151'];
            return `<span class="status-badge" style="background:#${{bg}};color:#${{fg}}">${{s}}</span>`;
          }}).join('');
          tr.appendChild(td);return;
        }}
      }}
      else if(key==='fileLocation'){{
        const url=row[key]||'';
        if(url){{td.innerHTML=`<a class="file-link" href="${{url}}" target="_blank">View File</a>`;tr.appendChild(td);return;}}
      }}
      else{{val=String(row[key]||'');}}
      td.textContent=val; tr.appendChild(td);
    }});
    const tdAct=document.createElement('td'); tdAct.className='actions-cell';
    if(CAN_EDIT) tdAct.innerHTML=`<button class="act-btn" onclick="editRecord('${{row._id}}')">✏</button> <button class="act-btn del" onclick="deleteRecord('${{row._id}}')">🗑</button>`;
    else tdAct.innerHTML='<span style="color:var(--muted);font-size:10px">—</span>';
    tr.appendChild(tdAct); body.appendChild(tr);
    if(!row._isRev) sr++;
  }});
  const ov=state.allRecords.filter(r=>r._overdue).length;
  document.getElementById('s-total').textContent='Total: '+state.allRecords.length;
  document.getElementById('s-showing').textContent='Showing: '+rows.length;
  document.getElementById('s-overdue').textContent='Overdue: '+ov;
}}

function calcWorkingDays(start,end){{
  if(!start||!end) return '';
  try{{
    let s=new Date(start),e=new Date(end);
    if(isNaN(s)||isNaN(e)||e<=s) return '0';
    let count=0,cur=new Date(s); cur.setDate(cur.getDate()+1);
    while(cur<=e){{if(cur.getDay()!==5) count++; cur.setDate(cur.getDate()+1);}}
    return String(count);
  }}catch(e){{return '';}}
}}

function sortBy(key){{
  state.sortDir=state.sortCol===key?(state.sortDir==='asc'?'desc':'asc'):'asc';
  state.sortCol=key; buildTableHead(); renderRows();
}}

function initColResize(){{
  document.querySelectorAll('#reg-table thead tr:first-child th[data-key]').forEach(th=>{{
    if(th.querySelector('.col-resizer')) return;
    const rz=document.createElement('div'); rz.className='col-resizer'; th.appendChild(rz);
    let startX,startW,dragging=false;
    rz.onmousedown=e=>{{e.stopPropagation();e.preventDefault();startX=e.clientX;startW=th.getBoundingClientRect().width;dragging=true;rz.classList.add('resizing');document.body.style.cursor='col-resize';document.body.style.userSelect='none';}};
    document.addEventListener('mousemove',e=>{{if(!dragging)return;const w=Math.max(40,startW+e.clientX-startX);th.style.width=w+'px';th.style.minWidth=w+'px';th.style.maxWidth=w+'px';}});
    document.addEventListener('mouseup',e=>{{if(!dragging)return;dragging=false;rz.classList.remove('resizing');document.body.style.cursor='';document.body.style.userSelect='';}});
  }});
}}

let _searchTimer;
function applySearch(){{clearTimeout(_searchTimer);_searchTimer=setTimeout(()=>loadRecords(),250);}}

// ── Add/Edit Record ──
function openAddRecord(){{
  state.editingId=null;
  document.getElementById('rec-modal-title').textContent='Add Document';
  document.getElementById('seq-warn').style.display='none';
  buildRecordForm(null); openModal('rec-modal');
}}
function editRecord(id){{
  state.editingId=id;
  const row=state.allRecords.find(r=>r._id===id); if(!row) return;
  document.getElementById('rec-modal-title').textContent='Edit Document';
  buildRecordForm(row); openModal('rec-modal');
}}
async function buildRecordForm(row){{
  const allCols=await apiP('/api/columns&dt='+state.activeTab);
  if(!allCols) return;
  const AUTO=new Set(['expectedReplyDate','duration']);
  const dt=state.docTypes.find(d=>d.id===state.activeTab);
  const prefix=dt?.code||state.activeTab;
  const grid=document.getElementById('rec-form'); grid.innerHTML='';
  let nextDocNo='';
  if(!row){{const r=await apiP('/api/next_doc_no/'+state.activeTab);nextDocNo=r?.next||'';}}
  for(const col of allCols){{
    if(AUTO.has(col.col_key)) continue;
    const key=col.col_key;
    const full=['title','remarks','fileLocation','itemRef'].includes(key);
    const grp=document.createElement('div'); grp.className='form-group'+(full?' full':'');
    const lbl=document.createElement('label'); lbl.textContent=col.label; grp.appendChild(lbl);
    const existing=row?.[key]||'';
    if(col.col_type==='date'){{
      const inp=document.createElement('input');inp.type='date';inp.id='f-'+key;inp.value=existing;grp.appendChild(inp);
    }}else if(col.col_type==='dropdown'&&col.list_name){{
      grp.appendChild(buildMultiSelect(key,state.allLists[col.list_name]||[],existing));
    }}else if(col.col_type==='docno'){{
      const inp=document.createElement('input');inp.id='f-'+key;inp.value=row?existing:nextDocNo;
      inp.style.fontFamily='Consolas,monospace';inp.style.fontWeight='600';
      inp.onblur=()=>checkSeqGap(inp.value,prefix);grp.appendChild(inp);
    }}else if(key==='remarks'){{
      const ta=document.createElement('textarea');ta.id='f-'+key;ta.value=existing;grp.appendChild(ta);
    }}else{{
      const inp=document.createElement('input');inp.id='f-'+key;inp.value=existing;
      if(col.col_type==='link') inp.placeholder='https://drive.google.com/...';
      grp.appendChild(inp);
    }}
    grid.appendChild(grp);
  }}
}}
function buildMultiSelect(key,options,initialValue){{
  const selected=initialValue?initialValue.split(',').map(s=>s.trim()).filter(Boolean):[];
  const container=document.createElement('div');container.className='ms-container';container.id='f-'+key;container.dataset.value=initialValue||'';
  function render(){{
    container.innerHTML='';
    selected.forEach(v=>{{const tag=document.createElement('span');tag.className='ms-tag';tag.innerHTML=`${{v}} <span class="rm" data-v="${{v}}">✕</span>`;tag.querySelector('.rm').onclick=e=>{{e.stopPropagation();selected.splice(selected.indexOf(v),1);container.dataset.value=selected.join(', ');render();}};container.appendChild(tag);}});
    if(!selected.length) container.innerHTML='<span class="ms-placeholder">Select...</span>';
    container.dataset.value=selected.join(', ');
  }}
  container.onclick=e=>{{
    if(e.target.classList.contains('rm')) return;
    const existing=document.querySelector('.ms-dropdown');if(existing){{existing.remove();return;}}
    const dd=document.createElement('div');dd.className='ms-dropdown';
    options.forEach(opt=>{{const item=document.createElement('div');item.className='ms-option'+(selected.includes(opt)?' sel':'');item.innerHTML=`<input type="checkbox" ${{selected.includes(opt)?'checked':''}} style="pointer-events:none"> ${{opt}}`;item.onclick=ev=>{{ev.stopPropagation();if(selected.includes(opt))selected.splice(selected.indexOf(opt),1);else selected.push(opt);container.dataset.value=selected.join(', ');render();item.classList.toggle('sel',selected.includes(opt));item.querySelector('input').checked=selected.includes(opt);}};dd.appendChild(item);}});
    container.style.position='relative';container.appendChild(dd);
  }};
  document.addEventListener('click',e=>{{if(!container.contains(e.target))container.querySelector('.ms-dropdown')?.remove();}},true);
  render();return container;
}}
function checkSeqGap(docNo,prefix){{
  const m=docNo.match(new RegExp(`^${{prefix}}-(\\\\d+)\\\\s+REV(\\\\d+)$`,'i'));
  if(!m||parseInt(m[2])!==0||parseInt(m[1])<=1){{document.getElementById('seq-warn').style.display='none';return;}}
  const prevNo=`${{prefix}}-${{String(parseInt(m[1])-1).padStart(3,'0')}} REV00`;
  const exists=state.allRecords.some(r=>r.docNo===prevNo);
  document.getElementById('seq-warn').style.display=exists?'none':'block';
}}
async function saveRecord(){{
  const allCols=await apiP('/api/columns&dt='+state.activeTab); if(!allCols) return;
  const AUTO=new Set(['expectedReplyDate','duration']);
  const data={{}};
  for(const col of allCols){{
    if(AUTO.has(col.col_key)) continue;
    const el=document.getElementById('f-'+col.col_key); if(!el) continue;
    data[col.col_key]=el.classList.contains('ms-container')?el.dataset.value||'':el.tagName==='TEXTAREA'?el.value.trim():el.value.trim();
  }}
  if(!data.docNo){{toast('Document No. is required','error');return;}}
  if(state.editingId) data._id=state.editingId;
  await apiP('/api/records/'+state.activeTab,{{method:'POST',body:JSON.stringify(data)}});
  closeModal('rec-modal'); await loadRecords(); await loadDocTypes();
  toast(state.editingId?'Record updated':'Record added','success');
}}
async function deleteRecord(id){{
  if(!confirm('Delete this record?')) return;
  await apiP('/api/delete_record/'+id,{{method:'POST',body:'{{}}'}});
  await loadRecords(); await loadDocTypes(); toast('Record deleted','warning');
}}

// ── Bulk delete ──
function updateBulkBar(){{
  const checked=document.querySelectorAll('.chk-cell input[data-id]:checked');
  document.getElementById('bulk-count').textContent=checked.length+' selected';
  document.getElementById('bulk-bar').classList.toggle('show',checked.length>0);
  const all=document.querySelectorAll('.chk-cell input[data-id]');
  const ca=document.getElementById('chk-all');
  if(ca) ca.checked=all.length>0&&checked.length===all.length;
}}
function toggleSelectAll(v){{document.querySelectorAll('.chk-cell input[data-id]').forEach(cb=>cb.checked=v);updateBulkBar();}}
function clearSelection(){{document.querySelectorAll('.chk-cell input').forEach(cb=>cb.checked=false);updateBulkBar();}}
async function bulkDelete(){{
  const ids=[...document.querySelectorAll('.chk-cell input[data-id]:checked')].map(cb=>cb.dataset.id);
  if(!ids.length) return;
  if(!confirm('Delete '+ids.length+' documents?')) return;
  let ok=0;
  for(const id of ids){{const r=await apiP('/api/delete_record/'+id,{{method:'POST'}});if(r&&r.ok)ok++;}}
  clearSelection(); await loadRecords(); await loadDocTypes();
  toast('✔ Deleted '+ok+' documents','success');
}}

// ── Project Modal ──
const PROJ_FIELDS=[['code','Code'],['name','Project Name'],['startDate','Start Date'],['endDate','End Date'],['client','Client'],['landlord','Landlord'],['pmo','PMO'],['mainConsultant','Main Consultant'],['mepConsultant','MEP Consultant'],['contractor','Contractor']];
async function openProjectModal(){{
  const proj=await apiP('/api/project'); if(!proj) return;
  let extra=[]; try{{extra=JSON.parse(proj.extraFields||'[]');}}catch(e){{}}
  const body=document.getElementById('proj-modal-body'); body.innerHTML='';
  const grid=document.createElement('div'); grid.className='form-grid';
  PROJ_FIELDS.forEach(([key,lbl])=>{{
    const grp=document.createElement('div');grp.className='form-group';
    const label=document.createElement('label');label.textContent=lbl;
    const inp=document.createElement('input');inp.id='pf-'+key;inp.value=proj[key]||'';
    inp.style.cssText='padding:7px 10px;border:1px solid var(--border);border-radius:var(--radius);font-family:inherit;font-size:12px;outline:none;width:100%';
    grp.appendChild(label);grp.appendChild(inp);grid.appendChild(grp);
  }});
  body.appendChild(grid);
  // Logos
  const logoTitle=document.createElement('div');logoTitle.className='section-title';logoTitle.textContent='Company Logos';body.appendChild(logoTitle);
  const logoGrid=document.createElement('div');logoGrid.style.cssText='display:grid;grid-template-columns:1fr 1fr;gap:12px';
  for(const [logoKey,logoLabel] of [['logo_left','Left Logo'],['logo_right','Right Logo']]){{
    const div=document.createElement('div');div.style.cssText='border:1px solid var(--border);border-radius:6px;padding:10px;text-align:center;background:var(--bg)';
    const lbl=document.createElement('div');lbl.style.cssText='font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;margin-bottom:6px';lbl.textContent=logoLabel;div.appendChild(lbl);
    const preview=document.createElement('img');preview.id='lp-'+logoKey;preview.style.cssText='max-height:52px;max-width:100%;object-fit:contain;display:block;margin:0 auto 6px';div.appendChild(preview);
    fetch('/api/logo/'+logoKey+'?p='+PID).then(r=>r.ok?r.blob():null).then(b=>{{if(b)preview.src=URL.createObjectURL(b);}});
    const fi=document.createElement('input');fi.type='file';fi.accept='image/*';fi.style.cssText='width:100%;font-size:10px;margin-bottom:4px';
    fi.onchange=async e=>{{const f=e.target.files[0];if(!f)return;const b64=await new Promise(res=>{{const fr=new FileReader();fr.onload=e2=>res(e2.target.result);fr.readAsDataURL(f);}});preview.src=b64;await apiP('/api/logo',{{method:'POST',body:JSON.stringify({{key:logoKey,data:b64.split(',')[1]}})}});toast('Logo saved','success');}};
    div.appendChild(fi);
    const clr=document.createElement('button');clr.className='btn btn-secondary btn-sm';clr.textContent='Remove';clr.style.fontSize='9px';
    clr.onclick=async()=>{{await apiP('/api/logo',{{method:'POST',body:JSON.stringify({{key:logoKey,data:''}})}});preview.src='';toast('Logo removed','warning');}};
    div.appendChild(clr);logoGrid.appendChild(div);
  }}
  body.appendChild(logoGrid);
  openModal('proj-modal');
}}
async function saveProject(){{
  const data={{}};
  PROJ_FIELDS.forEach(([key])=>{{const el=document.getElementById('pf-'+key);if(el)data[key]=el.value.trim();}});
  const r=await apiP('/api/project',{{method:'POST',body:JSON.stringify(data)}});
  if(r===null) return;
  if(r&&r.ok){{closeModal('proj-modal');toast('✔ Project saved!','success');
    PROJ_FIELDS.forEach(([key])=>{{const el=document.querySelector(`#projbar .pf-val[data-key="${{key}}"]`);if(el)el.textContent=data[key]||'—';}});
  }}else toast('Save failed','error');
}}

// ── Settings (dropdown lists) ──
async function openSettings(){{
  await loadLists(true);
  const body=document.getElementById('settings-body');body.innerHTML='';
  for(const [listName,items] of Object.entries(state.allLists)){{
    const title=document.createElement('div');title.className='section-title';title.textContent=listName.charAt(0).toUpperCase()+listName.slice(1);body.appendChild(title);
    const ul=document.createElement('ul');ul.className='settings-list';
    items.forEach(item=>{{const li=document.createElement('li');li.className='settings-item';li.innerHTML=`<span class="name">${{item}}</span><button onclick="removeListItem('${{listName}}','${{item.replace(/'/g,"\\\\'")}}'  ,this)">Remove</button>`;ul.appendChild(li);}});
    body.appendChild(ul);
    const addRow=document.createElement('div');addRow.className='add-item-row';
    addRow.innerHTML=`<input id="new-${{listName}}" placeholder="New item..."><button class="btn btn-success btn-sm" onclick="addListItem('${{listName}}')">Add</button>`;
    body.appendChild(addRow);
  }}
  body.innerHTML+=`<div class="section-title">Add New List</div><div class="add-item-row"><input id="new-list-name" placeholder="List name"><button class="btn btn-primary btn-sm" onclick="createNewList()">Create</button></div>`;
  openModal('settings-modal');
}}
async function addListItem(listName){{
  const inp=document.getElementById('new-'+listName);const val=inp?.value.trim();if(!val)return;
  await apiP('/api/dropdown_lists/add',{{method:'POST',body:JSON.stringify({{list_name:listName,item:val}})}});
  await loadLists(true);openSettings();
}}
async function removeListItem(listName,item,btn){{
  await apiP('/api/dropdown_lists/remove',{{method:'POST',body:JSON.stringify({{list_name:listName,item:item}})}});
  btn.closest('li').remove();await loadLists(true);
}}
async function createNewList(){{
  const name=document.getElementById('new-list-name')?.value.trim().toLowerCase().replace(/\\s+/g,'_');if(!name)return;
  await apiP('/api/dropdown_lists/add',{{method:'POST',body:JSON.stringify({{list_name:name,item:'Item 1'}})}});
  await loadLists(true);openSettings();
}}

// ── Doc Type ──
function openAddDocType(){{document.getElementById('dt-name').value='';document.getElementById('dt-code').value='';openModal('addtype-modal');}}
async function saveDocType(){{
  const name=document.getElementById('dt-name').value.trim();const code=document.getElementById('dt-code').value.trim().toUpperCase();
  if(!name||!code){{toast('Name and code required','error');return;}}
  await apiP('/api/doc_types',{{method:'POST',body:JSON.stringify({{name,code}})}});
  closeModal('addtype-modal');await loadDocTypes();switchTab(code);toast('Document type added','success');
}}

// ── Columns ──
async function manageColumns(){{
  const cols=await apiP('/api/columns&dt='+state.activeTab);if(!cols)return;
  const body=document.getElementById('cols-body');body.innerHTML='';
  const PROTECTED=new Set(['docNo','issuedDate','status','expectedReplyDate']);
  const ul=document.createElement('ul');ul.className='settings-list';ul.style.maxHeight='360px';
  cols.forEach(col=>{{
    const li=document.createElement('li');li.className='settings-item';
    li.innerHTML=`<input type="checkbox" ${{col.visible?'checked':''}} onchange="toggleColVis(${{col.id}},this.checked)"><span class="name">${{col.label}}</span><span style="font-size:9px;background:#e0e7ff;color:#3730a3;padding:1px 6px;border-radius:3px">${{col.col_type}}</span>${{!PROTECTED.has(col.col_key)?`<button onclick="delCol(${{col.id}},this)">Del</button>`:'<span style="font-size:9px;color:var(--muted)">core</span>'}}`;
    ul.appendChild(li);
  }});
  body.appendChild(ul);openModal('cols-modal');
}}
async function toggleColVis(colId,visible){{await apiP('/api/columns/visibility/'+colId,{{method:'POST',body:JSON.stringify({{visible}})}});}}
async function delCol(colId,btn){{if(!confirm('Delete column?'))return;await apiP('/api/columns/delete/'+colId,{{method:'POST',body:'{{}}'}});btn.closest('li').remove();}}

function onColTypeChange(val){{
  document.getElementById('col-list-grp').style.display=val==='dropdown'?'flex':'none';
  document.getElementById('col-dur-grp').style.display=val==='duration_calc'?'flex':'none';
  document.getElementById('col-dur-end-grp').style.display=val==='duration_calc'?'flex':'none';
  document.getElementById('col-dur-info').style.display=val==='duration_calc'?'block':'none';
}}
async function openAddColumn(){{
  await loadLists();
  const sel=document.getElementById('col-list-src');
  sel.innerHTML=Object.keys(state.allLists).map(k=>`<option value="${{k}}">${{k}}</option>`).join('');
  const allCols=await apiP('/api/columns&dt='+state.activeTab);
  const dateCols=(allCols||[]).filter(c=>['date','auto_date'].includes(c.col_type));
  const dateOpts=dateCols.map(c=>`<option value="${{c.col_key}}">${{c.label}}</option>`).join('')||'<option value="issuedDate">Issued Date</option>';
  document.getElementById('col-dur-start-sel').innerHTML=dateOpts;
  document.getElementById('col-dur-end-sel').innerHTML=dateOpts;
  document.getElementById('col-name').value='';document.getElementById('col-type').value='text';onColTypeChange('text');
  openModal('addcol-modal');
}}
async function saveAddColumn(){{
  const name=document.getElementById('col-name').value.trim();
  const type=document.getElementById('col-type').value;
  const listSrc=type==='dropdown'?document.getElementById('col-list-src').value:null;
  const durStart=type==='duration_calc'?document.getElementById('col-dur-start-sel').value:null;
  const durEnd=type==='duration_calc'?document.getElementById('col-dur-end-sel').value:null;
  if(!name){{toast('Column name required','error');return;}}
  if(type==='duration_calc'&&durStart===durEnd){{toast('Start and end must be different','error');return;}}
  const key='custom_'+name.toLowerCase().replace(/[^a-z0-9]+/g,'_')+'_'+Date.now();
  const listName=type==='duration_calc'?(durStart+','+durEnd):listSrc;
  await apiP('/api/columns/add',{{method:'POST',body:JSON.stringify({{dt_id:state.activeTab,col_key:key,label:name,col_type:type,list_name:listName}})}});
  closeModal('addcol-modal');closeModal('cols-modal');await loadRecords();toast('✔ Column added','success');
}}

// ── Export / Import ──
function doExport(){{if(!state.activeTab)return;window.location='/api/export/'+state.activeTab+'?p='+PID;}}
function openImport(){{openModal('import-modal');}}
async function doImport(){{
  const file=document.getElementById('import-file').files[0];if(!file)return;
  const btn=document.querySelector('#import-modal .btn-primary');
  if(btn){{btn.disabled=true;btn.textContent='⏳ Importing...';}}
  try{{
    const ext=file.name.split('.').pop().toLowerCase();
    if(ext==='xlsx'||ext==='xls'){{
      const b64=await new Promise((res,rej)=>{{const fr=new FileReader();fr.onload=e=>res(e.target.result);fr.onerror=rej;fr.readAsDataURL(file);}});
      const r=await apiP('/api/import_xlsx',{{method:'POST',body:JSON.stringify({{dt_id:state.activeTab,file_b64:b64}})}});
      if(r===null)return;closeModal('import-modal');switchTab(state.activeTab);await loadDocTypes();toast('✔ Imported '+r.imported+' records','success');
    }}else if(ext==='csv'){{
      const text=await file.text();
      const r=await apiP('/api/import_csv',{{method:'POST',body:JSON.stringify({{dt_id:state.activeTab,csv_text:text}})}});
      if(r===null)return;closeModal('import-modal');switchTab(state.activeTab);await loadDocTypes();toast('✔ Imported '+r.imported+' records','success');
    }}else{{toast('Only .csv or .xlsx supported','error');}}
  }}catch(e){{toast('Import error: '+e.message,'error');}}
  finally{{if(btn){{btn.disabled=false;btn.textContent='Import';}}}}
}}

init();
</script></body></html>"""


# ── SERVER START ──────────────────────────────────────────────
if __name__ == "__main__":
    import threading

    db.init_db()

    local_ip = "127.0.0.1"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]; s.close()
    except: pass

    if IS_RENDER:
        print(f"[DCR v6] Starting on Render — PORT={PORT}")
        print(f"[DCR v6] DB: {'PostgreSQL/Supabase' if db.USE_POSTGRES else 'SQLite'}")
    else:
        print("=" * 52)
        print("  📋 Document Control Register v6")
        print("=" * 52)
        print(f"  ✅ Local:   http://localhost:{PORT}")
        print(f"  🌐 Network: http://{local_ip}:{PORT}")
        print("=" * 52)
        def open_browser():
            import time, webbrowser; time.sleep(1)
            webbrowser.open(f"http://localhost:{PORT}")
        threading.Thread(target=open_browser, daemon=True).start()

    db.cleanup_sessions()
    server = HTTPServer(("0.0.0.0", PORT), DCRHandler)
    try: server.serve_forever()
    except KeyboardInterrupt: print("\nServer stopped.")
