"""blueprints/exporting.py - DCR Export/Import API routes.

Handles export/import endpoints:
  GET  /api/export_all/<pid>
  GET  /api/export/<pid>/<dt_id>
  helper: _build_pdf_for_dt(pid, dt_id, proj, buf=None)
  GET  /api/export_pdf/<pid>/<dt_id>
  GET  /api/export_pdf_all/<pid>
  POST /api/import/<pid>/<dt_id>

Step 13 of the incremental refactor.
Logic is identical to the original app.py routes/helper — only the decorator
changed from @app.route to @exporting_bp.route.
"""
import base64
import datetime
import io
import logging
import re
import uuid

from flask import Blueprint, jsonify, request, send_file

import db
from auth import current_user, can_edit
from utils import compute_expected_reply, compute_duration, is_overdue, format_date, extract_rev

exporting_bp = Blueprint("exporting", __name__)
logger = logging.getLogger(__name__)


def _normalize_sheet_name(name):
    s = str(name or "").strip().lower()
    s = re.sub(r"^\s*\d+\s*[\.\-)\]]*\s*", "", s)
    s = re.sub(r"[(){}\[\],.:;_]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" -_/")
    return s


def _sheet_aliases(dt):
    vals = {
        str(dt.get("id", "")).strip().lower(),
        str(dt.get("code", "")).strip().lower(),
        str(dt.get("name", "")).strip().lower(),
        _normalize_sheet_name(dt.get("id", "")),
        _normalize_sheet_name(dt.get("code", "")),
        _normalize_sheet_name(dt.get("name", "")),
    }
    code = str(dt.get("code", "")).strip().lower()
    name = _normalize_sheet_name(dt.get("name", ""))
    if code == "ltr" or "letter" in name:
        vals.update({"letter", "letters"})
    if code == "pr" or "requisition" in name:
        vals.update({"purchase requisition", "purchase requisitions"})
    return {v for v in vals if v}


def _match_sheet_to_dt(sheet_name, dts):
    raw = str(sheet_name or "").strip()
    low = raw.lower()
    norm = _normalize_sheet_name(raw)
    no_paren = _normalize_sheet_name(re.sub(r"\([^)]*\)", " ", raw))
    no_prefix = _normalize_sheet_name(re.sub(r"^\s*\d+\s*[\.\-)\]]*\s*", "", raw))
    paren_code = ""
    m = re.search(r"\(([A-Za-z0-9_-]+)\)", raw)
    if m:
        paren_code = m.group(1).strip().lower()

    def exact_match(field):
        for dt in dts:
            if str(dt.get(field, "")).strip().lower() == low:
                return dt
        return None

    for field in ("id", "code", "name"):
        dt = exact_match(field)
        if dt:
            return dt

    for dt in dts:
        aliases = _sheet_aliases(dt)
        if norm in aliases or no_paren in aliases or no_prefix in aliases:
            return dt

    if paren_code:
        for field in ("id", "code"):
            for dt in dts:
                if str(dt.get(field, "")).strip().lower() == paren_code:
                    return dt

    return None


def _has_meaningful_values(row_data):
    for v in row_data.values():
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip():
                return True
        elif str(v).strip():
            return True
    return False


def _normalize_match_value(value):
    return str(value or "").strip().lower()


def _save_import_row(pid, dt_id, row_data):
    doc_no = str(row_data.get("docNo", "") or "").strip()
    existing = db.get_record_by_doc_no(pid, dt_id, doc_no) if _normalize_match_value(doc_no) else None
    if existing:
        merged = db.merge_record_data(existing, row_data)
        db.save_record(pid, dt_id, existing["_id"], merged)
        return "updated"
    db.save_record(pid, dt_id, str(uuid.uuid4()), row_data)
    return "created"


def _import_excel_worksheet(pid, dt_id, ws, cols):
    import datetime as _dt
    import re

    def _norm(s):
        return re.sub(r'[^a-z0-9]', '', str(s).lower()) if s else ""

    col_map = {}
    for c in cols:
        col_map[str(c.get("label", "")).strip()] = c["col_key"]
        col_map[_norm(c.get("label", ""))] = c["col_key"]
        col_map[_norm(c.get("col_key", ""))] = c["col_key"]

    header = None
    imported = 0
    created = 0
    updated = 0
    skipped_blank = 0
    skipped_invalid = 0
    warnings = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = []
        for v in row:
            if v is None: vals.append("")
            elif isinstance(v, (_dt.datetime, _dt.date)):
                vals.append((v.date() if isinstance(v, _dt.datetime) else v).strftime("%Y-%m-%d"))
            else: vals.append(str(v).strip())
        if not any(vals):
            skipped_blank += 1
            continue
        if header is None:
            # Strict header detection to skip noisy/summary rows
            is_header = False
            matched_cols = 0
            for v in vals:
                if not v: continue
                n_v = _norm(v)
                if n_v in ("sr", "documentno", "docno", "nocno", "irno", "mirno", "ncrno"):
                    is_header = True
                if str(v).strip() in col_map or n_v in col_map:
                    matched_cols += 1
            
            if is_header or matched_cols >= 3:
                header = []
                for v in vals:
                    v_str = str(v).strip() if v else ""
                    n_v = _norm(v_str)
                    if v_str in col_map:
                        header.append(col_map[v_str])
                    elif n_v in col_map:
                        header.append(col_map[n_v])
                    else:
                        header.append(v_str)
            continue
        
        row_data = {header[i]: v for i, v in enumerate(vals)
                    if i < len(header) and header[i] and header[i] not in ("Sr.","sr","")}
        if not _has_meaningful_values(row_data):
            skipped_blank += 1
            continue
            
        try:
            action = _save_import_row(pid, dt_id, row_data)
            imported += 1
            if action == "updated":
                updated += 1
            else:
                created += 1
        except Exception as e:
            skipped_invalid += 1
            logger.warning("import_row_skipped pid=%s dt_id=%s row=%s error=%s",
                           pid, dt_id, row_idx, e)
            if len(warnings) < 20:
                warnings.append({"row": row_idx, "error": str(e)})
    return imported, created, updated, header is not None, skipped_blank, skipped_invalid, warnings


def _is_pr_dt(dt):
    if not dt: return False
    code = str(dt.get("code","")).strip().upper()
    name = str(dt.get("name","")).strip().lower()
    dt_id = str(dt.get("id","")).strip().upper()
    return dt_id == "PR" or code == "PR" or "requisition" in name or "purchase request" in name


def _pr_details_key(cols):
    def _norm(v):
        return re.sub(r"[^a-z0-9]+", "", str(v or "").strip().lower())
    for c in cols:
        key = str(c.get("col_key", ""))
        label = str(c.get("label", "")).strip()
        if label == "PR Details" or key in ("prDetails", "pr_details"):
            return c.get("col_key")
    for c in cols:
        nk = _norm(c.get("col_key", ""))
        nl = _norm(c.get("label", ""))
        if nk == "prdetails" or nl == "prdetails":
            return c.get("col_key")
    for c in cols:
        nk = _norm(c.get("col_key", ""))
        nl = _norm(c.get("label", ""))
        if "pr" in nk and "detail" in nk:
            return c.get("col_key")
        if "pr" in nl and "detail" in nl:
            return c.get("col_key")
    return None


def _pr_items_text(items):
    if not items: return ""
    lines = []
    for it in items:
        row_type = str(it.get("row_type","item") or "item").strip().lower()
        name = str(it.get("item_name","") or "").strip()
        if row_type == "header":
            if name:
                lines.append(name)
            continue
        unit = str(it.get("unit","") or "").strip()
        qty  = it.get("quantity")
        remarks = str(it.get("remarks","") or "").strip()
        if qty is not None and qty != "":
            qv = str(qty)
        else:
            qv = ""
        parts = [p for p in [name, unit, qv] if p]
        line = " | ".join(parts) if parts else ""
        if remarks:
            line = (line + " — " if line else "") + remarks
        if line:
            lines.append(line)
    return "\n".join(lines)


def _resolve_pr_details_value(row, pr_items_map, pr_details_key):
    manual = str(row.get(pr_details_key, "") or "").strip() if pr_details_key else ""
    if manual:
        return manual
    return _pr_items_text(pr_items_map.get(row.get("_id"), []))


def _safe_excel_name_part(val, fallback):
    text = re.sub(r'[\\/:*?"<>|]+', "_", str(val or "").strip())
    text = re.sub(r"\s+", "_", text).strip("._")
    return text or fallback


def _pick_first(row, keys):
    for key in keys:
        val = str(row.get(key, "") or "").strip()
        if val:
            return val
    return ""


def _is_floor_field(col_key, label=""):
    key = str(col_key or "").strip().lower()
    lbl = str(label or "").strip().lower()
    return key == "floor" or lbl in ("floor", "floors")


def _is_item_ref_field(col_key, label=""):
    key = str(col_key or "").strip().lower()
    lbl = re.sub(r"[_./-]+", " ", str(label or "").strip().lower())
    return key == "itemref" or ("item ref" in lbl and "dwg" in lbl)


def _format_multiline_display_value(col_key, label, value):
    text = str(value or "")
    if not text:
        return ""
    if _is_floor_field(col_key, label):
        return "\n".join([p.strip() for p in text.split(",") if p.strip()])
    if _is_item_ref_field(col_key, label):
        return text.replace("\r\n", "\n").replace("\r", "\n")
    return text


_REV_TOKEN_RE = re.compile(r"\bREV\s*[-_ ]?(\d+)\b", re.IGNORECASE)


def normalize_doc_base(doc_no):
    text = str(doc_no or "").strip()
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(?i)(?:[\s/_-]+)?REV\s*[-_ ]?\d+\s*$", "", text).strip()
    text = re.sub(r"(?i)\s+REV\s*[-_ ]?\d+\b", " ", text).strip()
    return re.sub(r"\s+", " ", text) or str(doc_no or "").strip()


def extract_rev_no(doc_no):
    match = _REV_TOKEN_RE.search(str(doc_no or ""))
    if not match:
        return 0
    try:
        return int(match.group(1))
    except ValueError:
        return 0


def _natural_sort_key(value):
    parts = re.split(r"(\d+)", str(value or ""))
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def _doc_revision_sort_key(row):
    doc_no = row.get("docNo", "") if isinstance(row, dict) else ""
    base = normalize_doc_base(doc_no)
    return (
        _natural_sort_key(base),
        extract_rev_no(doc_no),
        _natural_sort_key(doc_no),
        str(row.get("_id", "")) if isinstance(row, dict) else "",
    )


def _excel_column_width(col_key, label):
    key = str(col_key or "").strip()
    key_l = key.lower()
    label_l = re.sub(r"[_./-]+", " ", str(label or "").strip().lower())
    hay = f"{key_l} {label_l}"

    if key == "_sr":
        return 6
    if key_l in ("docno", "doc_no") or "document no" in hay or "letter ref" in hay:
        return 24
    if "title" in hay or "subject" in hay or "description" in hay:
        return 46
    if "pr detail" in hay or key_l in ("prdetails", "pr_details"):
        return 50
    if "remarks" in hay or "comment" in hay or "note" in hay:
        return 34
    if "ms ref" in hay or "dwg" in hay or "item ref" in hay or "reference" in hay:
        return 26
    if "file" in hay or "location" in hay or "link" in hay:
        return 22
    if "prepared" in hay or "company" in hay or "client" in hay or "consultant" in hay or "contractor" in hay:
        return 24
    if label_l in ("from", "to") or key_l in ("from", "to"):
        return 18
    if "discipline" in hay:
        return 16
    if "sub trade" in hay or "trade" in hay:
        return 18
    if "status" in hay:
        return 24
    if "date" in hay:
        return 14
    if "duration" in hay:
        return 10
    if "floor" in hay:
        return 18
    if "direction" in hay or "revision" in hay:
        return 12
    return 16


def _field_width_role(col):
    key = str(col.get("col_key", "") or "").strip()
    key_l = key.lower()
    label = str(col.get("label", "") or "").strip()
    label_l = re.sub(r"[_./-]+", " ", label.lower())
    hay = f"{key_l} {label_l}"

    if key == "_sr":
        return "serial"
    if key_l in ("docno", "doc_no") or "document no" in hay or "letter ref" in hay:
        return "document_no"
    if "pr detail" in hay or key_l in ("prdetails", "pr_details"):
        return "very_long"
    if "title" in hay or "subject" in hay:
        return "very_long"
    if "status" in hay:
        return "status"
    if "item description" in hay or "description" in hay or "scope" in hay:
        return "long"
    if "remarks" in hay or "comment" in hay or "note" in hay:
        return "long"
    if "file" in hay or "location" in hay or "link" in hay:
        return "file_link"
    if "ms ref" in hay or "dwg" in hay or "item ref" in hay or "parent letter" in hay or "reference" in hay:
        return "technical_ref"
    if "prepared" in hay or "engineer" in hay or "person" in hay:
        return "medium_wide"
    if label_l in ("from", "to") or key_l in ("from", "to"):
        return "medium"
    if "company" in hay or "client" in hay or "consultant" in hay or "contractor" in hay or "party" in hay:
        return "medium"
    if "discipline" in hay:
        return "medium"
    if "sub trade" in hay or "trade" in hay or "brand" in hay or "floor" in hay or "level" in hay:
        return "medium"
    if "duration" in hay:
        return "compact"
    if "date" in hay or "issued" in hay or "submitted" in hay or "received" in hay or "reply" in hay:
        return "date"
    if "direction" in hay or "revision" in hay or re.search(r"\brev\b", hay) or "code" in hay:
        return "compact"
    if "qty" in hay or "quantity" in hay or "unit" in hay:
        return "compact"
    return "default"


def _excel_width_from_role(role):
    return {
        "serial": 5.5,
        "document_no": 30,
        "very_long": 58,
        "long": 42,
        "technical_ref": 31,
        "status": 27,
        "file_link": 17,
        "medium_wide": 24,
        "medium": 18,
        "date": 14,
        "compact": 9.5,
        "default": 16,
    }.get(role, 16)


def _excel_width_from_web_px(px, role):
    try:
        px = float(px)
    except (TypeError, ValueError):
        return None
    if px <= 0:
        return None
    min_by_role = {
        "serial": 5,
        "document_no": 28,
        "very_long": 48,
        "long": 28,
        "technical_ref": 26,
        "status": 25,
        "file_link": 15,
        "medium_wide": 22,
        "medium": 15,
        "date": 13,
        "compact": 8.5,
        "default": 13,
    }
    max_by_role = {
        "serial": 8,
        "document_no": 36,
        "very_long": 72,
        "long": 54,
        "technical_ref": 40,
        "status": 32,
        "file_link": 22,
        "medium_wide": 34,
        "medium": 26,
        "date": 18,
        "compact": 14,
        "default": 24,
    }
    width = px / 7.2
    return max(min_by_role.get(role, 12), min(max_by_role.get(role, 28), width))


def _excel_sheet_column_widths(cols, dt_name=None, web_widths=None):
    """Build the Excel width profile from the actual ordered visible register columns."""
    web_widths = web_widths or {}
    all_cols = [{"col_key":"_sr","label":"Sr."}] + [
        {"col_key":c["col_key"],"label":c["label"]} for c in cols
    ]
    roles = [_field_width_role(c) for c in all_cols]
    widths = []
    for col, role in zip(all_cols, roles):
        live_width = _excel_width_from_web_px(web_widths.get(col["col_key"]), role)
        widths.append(live_width if live_width is not None else _excel_width_from_role(role))

    # Keep dense sheets readable without flattening every register into one generic profile.
    dt_hint = str(dt_name or "").lower()
    visible_keys = [str(c.get("col_key", "") or "").lower() for c in cols]
    visible_labels = [str(c.get("label", "") or "").lower() for c in cols]
    is_letters = (
        "letter" in dt_hint
        or any(k in {"from", "to", "fromparty", "toparty", "direction"} for k in visible_keys)
        and any("subject" in l or "letter" in l for l in visible_labels)
    )
    is_pr = (
        "purchase" in dt_hint
        or "requisition" in dt_hint
        or any("pr detail" in l or k in {"prdetails", "pr_details"} for k, l in zip(visible_keys, visible_labels))
    )

    if is_letters:
        for i, col in enumerate(all_cols):
            key = str(col.get("col_key", "") or "").lower()
            label = str(col.get("label", "") or "").lower()
            if "subject" in label:
                widths[i] = max(widths[i], 58)
            elif "letter ref" in label or key in {"letterref", "letter_ref", "docno"}:
                widths[i] = max(widths[i], 31)
            elif key in {"from", "to", "fromparty", "toparty"}:
                widths[i] = max(widths[i], 22)
            elif "parent letter" in label:
                widths[i] = max(widths[i], 30)
            elif "direction" in label or key == "direction":
                widths[i] = min(widths[i], 12)
    if is_pr:
        for i, col in enumerate(all_cols):
            label = str(col.get("label", "") or "").lower()
            key = str(col.get("col_key", "") or "").lower()
            if "pr detail" in label or key in {"prdetails", "pr_details"}:
                widths[i] = max(widths[i], 60)
            elif "ms ref" in label:
                widths[i] = max(widths[i], 28)

    # Add safety padding for Auto-fit
    for i in range(len(widths)):
        widths[i] += 10

    return widths


def _excel_col_hay(col_key, label):
    key = str(col_key or "").strip().lower()
    label_l = re.sub(r"[_./-]+", " ", str(label or "").strip().lower())
    return re.sub(r"\s+", " ", f"{key} {label_l}").strip()


def _is_excel_date_col(col_key, label):
    hay = _excel_col_hay(col_key, label)
    key = str(col_key or "").strip().lower()
    date_tokens = (
        "date", "issued", "submitted", "submission", "received", "reply",
        "expected", "actual", "rec from", "rec by", "received from", "received by", "sent"
    )
    if _is_excel_duration_col(col_key, label) or key == "_sr" or "prepared by" in hay or "reply status" in hay:
        return False
    return any(token in hay for token in date_tokens) or bool(re.search(r"\brec\b", hay))


def _is_excel_duration_col(col_key, label):
    key = str(col_key or "").strip().lower()
    hay = _excel_col_hay(col_key, label)
    return key == "duration" or "duration" in hay or re.search(r"\bdur\.?\b", hay) is not None


def _looks_like_excel_date_value(value):
    text = str(value or "").strip()
    if not text:
        return False
    text = text.replace("T", " ")
    text = re.sub(r"\s+\d{2}:\d{2}(?::\d{2})?(?:\.\d+)?$", "", text)
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}|\d{2}[-/.]\d{2}[-/.]\d{4}|[0-3]?\d[-/ ][A-Za-z]{3,9}[-/ ]\d{4}", text))


def _parse_excel_date_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("T", " ")
    text = re.sub(r"\s+\d{2}:\d{2}(?::\d{2})?(?:\.\d+)?$", "", text)
    text = re.sub(r"\s+", " ", text)
    for fmt in (
        "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
        "%d/%b/%Y", "%d-%b-%Y", "%d %b %Y", "%d/%B/%Y", "%d-%B-%Y",
    ):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _excel_cell_value(col_key, label, value):
    if _is_excel_date_col(col_key, label) or _looks_like_excel_date_value(value):
        parsed = _parse_excel_date_value(value)
        if parsed:
            return parsed
    return value


def _excel_center_col(col_key, label):
    key = str(col_key or "").strip()
    label_l = str(label or "").strip().lower()
    return (
        key in {"_sr", "issuedDate", "expectedReplyDate", "actualReplyDate"}
        or _is_excel_duration_col(col_key, label)
        or "date" in key.lower()
        or "date" in label_l
        or key == "status"
        or "status" in label_l
        or "docno" in key.lower()
        or "document" in label_l
        or "rev" in key.lower()
        or "revision" in label_l
        or "delay" in label_l
    )


def _excel_row_height(values, cols=None, widths=None, base=19):
    max_lines = 1
    cols = cols or []
    widths = widths or []
    for idx, value in enumerate(values):
        text = str(value or "")
        if not text:
            continue
        col = cols[idx] if idx < len(cols) else {}
        width = widths[idx] if idx < len(widths) else 18
        hay = _excel_col_hay(col.get("col_key", ""), col.get("label", ""))
        explicit = text.count("\n") + 1
        chars_per_line = max(14, int(width * 1.25))
        wrapped = max(1, (max((len(part) for part in text.split("\n")), default=0) // chars_per_line) + 1)
        line_cap = 2 if ("status" in hay or "document no" in hay or "letter ref" in hay) else 3
        if "title" in hay or "subject" in hay or "remarks" in hay or "description" in hay or "pr detail" in hay:
            line_cap = 4
        max_lines = max(max_lines, min(line_cap, max(explicit, wrapped)))
    return max(base, min(34, 13 * max_lines + 4))


def _excel_wrap_cell(col_key, label):
    return not (str(col_key or "") == "_sr" or _is_excel_duration_col(col_key, label) or _is_excel_date_col(col_key, label))


def _write_register_excel_sheet(ws, proj, dt, cols, records, pr_items_map=None, pr_details_key=None, web_widths=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    pr_items_map = pr_items_map or {}
    PRIMARY = "1F4E78"; PL = "2563A8"; WHITE = "FFFFFF"
    ALT = "F8FAFC"; OV = "FFF5F5"; MUTED = "9CA3AF"

    STATUS_XL = {
        "A - Approved":              ("C6EFCE","064E3B"),
        "B - Approved As Noted":     ("C6EFCE","064E3B"),
        "B,C - Approved & Resubmit": ("FFEB9C","713F12"),
        "C - Revise & Resubmit":     ("FFC7CE","991B1B"),
        "D - Review not Required":   ("FFC7CE","991B1B"),
        "Under Review":              ("FFEB9C","713F12"),
        "Cancelled":                 ("FFC7CE","991B1B"),
        "Open":                      ("FFEB9C","713F12"),
        "Closed":                    ("C6EFCE","064E3B"),
        "Replied":                   ("C6EFCE","064E3B"),
        "Pending":                   ("FFEB9C","713F12"),
    }

    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(): s=Side(style="thin",color="B2B2B2"); return Border(left=s,right=s,top=s,bottom=s)

    ws.sheet_view.showGridLines = False
    all_cols = [{"col_key":"_sr","label":"Sr."}] + [{"col_key":c["col_key"],"label":c["label"]} for c in cols]
    nc = len(all_cols)

    def mcell(row, val, bg, fg="FFFFFF", bold=False, sz=11, halign="center"):
        c = ws.cell(row=row, column=1, value=val)
        ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
        c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        return c

    dt_name = (dt.get("name") if dt else "") or (dt.get("id") if dt else "") or "REGISTER"
    pid = (proj or {}).get("id")
    dt_id = (dt or {}).get("id")
    expected_reply_rule = db.get_expected_reply_rule(pid, dt_id) if pid and dt_id else None
    
    brand_text = f"{proj.get('name', '')} - {proj.get('code', '')}".strip(" -")
    if not brand_text: brand_text = "DCR SYSTEM"
    mcell(1, brand_text, PRIMARY, bold=True, sz=14)
    ws.row_dimensions[1].height = 30

    info = f"DOCUMENT CONTROL REGISTER  -  {str(dt_name).upper()}   |   Exported: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    mcell(2, info, PL, sz=10)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 3

    ws.freeze_panes = "A5"
    ws.print_title_rows = "4:4"
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.left.text = "Generated from DCR System"

    col_widths = _excel_sheet_column_widths(cols, dt_name=dt_name, web_widths=web_widths)
    for ci, col in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
        c = ws.cell(row=4, column=ci, value=col["label"])
        c.font = Font(bold=True,color=WHITE,size=10,name="Arial")
        c.fill = fill(PRIMARY)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = thin()
    ws.row_dimensions[4].height = 22

    has_exp_col = any(c["col_key"]=="expectedReplyDate" for c in all_cols)
    sr = 1
    if not records:
        c = ws.cell(row=5, column=1, value="No records in this register")
        c.font = Font(italic=True, size=10, name="Arial", color=MUTED)
        ws.merge_cells(f"A5:{get_column_letter(nc)}5")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 24
        total_row = 6
    else:
        for ri, row in enumerate(records):
            rn   = 5 + ri
            is_rev = extract_rev(row.get("docNo","")) > 0
            status_val = row.get("status")
            action_val = row.get("action")
            ov     = is_overdue(row.get("issuedDate"), row.get("docNo"), row.get("actualReplyDate"), has_exp_col, expected_reply_rule, status=status_val, action=action_val)
            bg     = OV if ov else (ALT if sr%2==0 else WHITE)
            row_values = []

            for ci, col in enumerate(all_cols, 1):
                key = col["col_key"]
                if key=="_sr":                  val = "" if is_rev else str(sr)
                elif key=="expectedReplyDate":  val = format_date(compute_expected_reply(row.get("issuedDate"),row.get("docNo"), expected_reply_rule, status=status_val, action=action_val))
                elif _is_excel_duration_col(key, col["label"]):
                    dur_val = compute_duration(row.get("issuedDate"), row.get("actualReplyDate"), expected_reply_rule, status=status_val, action=action_val)
                    val = str(dur_val) if dur_val is not None else ""
                elif key=="issuedDate":         val = format_date(row.get(key,""))
                elif key=="actualReplyDate":    val = format_date(row.get(key,""))
                elif pr_details_key and key == pr_details_key:
                    val = _resolve_pr_details_value(row, pr_items_map, pr_details_key) or str(row.get(key,"") or "")
                else:                           val = str(row.get(key,"") or "")
                val = _format_multiline_display_value(key, col["label"], val)
                row_values.append(val)
                cell_value = _excel_cell_value(key, col["label"], val)

                if key == "fileLocation" and val and val.startswith("http"):
                    c = ws.cell(row=rn, column=ci)
                    c.value = "View"
                    c.hyperlink = val
                    c.font = Font(size=9,name="Arial",color="2563A8",underline="single")
                else:
                    c = ws.cell(row=rn, column=ci, value=cell_value)
                    if isinstance(cell_value, datetime.date):
                        c.number_format = "DD-MM-YYYY"
                    if _is_excel_duration_col(key, col["label"]) and val == "0":
                        c.value = 0
                    if key=="status" and val:
                        bg2, fg2 = STATUS_XL.get(val, ("F3F4F6","374151"))
                        c.fill = fill(bg2); c.font = Font(bold=True,size=9,name="Arial",color=fg2)
                    elif key=="docNo":
                        c.fill = fill(bg)
                        c.font = Font(size=10,name="Consolas",bold=not is_rev,
                                      color=MUTED if is_rev else PRIMARY)
                    else:
                        c.fill = fill(bg)
                        c.font = Font(size=10,name="Arial",
                                      color=MUTED if is_rev else ("991B1B" if ov else "1E2A3A"))
                c.border    = thin()
                c.alignment = Alignment(
                    vertical="center",
                    wrap_text=_excel_wrap_cell(key, col["label"]),
                    horizontal="center" if _excel_center_col(key, col["label"]) else "left",
                )
            ws.row_dimensions[rn].height = _excel_row_height(row_values, all_cols, col_widths, base=19)
            if not is_rev: sr += 1
        total_row = 5 + len(records)

    real = sum(1 for r in records if extract_rev(r.get("docNo",""))==0)
    mcell(total_row, f"TOTAL: {real} documents  |  {len(records)} submissions", PRIMARY, sz=10, halign="left")
    ws.row_dimensions[total_row].height = 22
    if total_row >= 4:
        ws.auto_filter.ref = f"A4:{get_column_letter(nc)}{max(4, total_row - 1)}"
        
    # TRUE AUTO-FIT
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            # Skip checking lengths for merged cells to avoid errors
            if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                continue
            try:
                if cell.row <= 3: continue
                if cell.value:
                    lines = str(cell.value).split('\n')
                    longest = max(len(l) for l in lines) if lines else 0
                    max_length = max(max_length, longest)
            except: pass
        adjusted = min(55, max_length + 5)
        ws.column_dimensions[col_letter].width = max(12, adjusted)



def _write_summary_dashboard(ws, proj, records_by_dt):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    PRIMARY = "1F4E78"; WHITE = "FFFFFF"; LIGHT = "F8FAFC"
    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(): s=Side(style="thin",color="B2B2B2"); return Border(left=s,right=s,top=s,bottom=s)
    
    ws.sheet_view.showGridLines = False
    for col, width in {"A":4, "B":30, "C":15, "D":15, "E":15}.items():
        ws.column_dimensions[col].width = width

    def mcell(row, val, bg, fg="FFFFFF", bold=False, sz=11, halign="center"):
        c = ws.cell(row=row, column=2, value=val)
        ws.merge_cells(f"B{row}:E{row}")
        c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        c.border = thin()
        return c

    mcell(2, f"EXECUTIVE SUMMARY DASHBOARD - {proj.get('name', 'PROJECT')}", PRIMARY, bold=True, sz=14)
    ws.row_dimensions[2].height = 30
    
    mcell(3, f"Project Code: {proj.get('code','')} | Date: {datetime.datetime.now().strftime('%d-%m-%Y')}", "2563A8", sz=10)
    ws.row_dimensions[3].height = 20

    row_idx = 5
    headers = ["Document Type", "Total Records", "Approved / Closed", "Pending / Overdue"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=row_idx, column=i, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = fill(PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    ws.row_dimensions[row_idx].height = 25
    row_idx += 1

    total_all = 0
    total_approved = 0
    total_pending = 0
    
    for dt_name, records in records_by_dt.items():
        count = len(records)
        if count == 0: continue
        total_all += count
        
        approved = sum(1 for r in records if str(r.get("status", "")).strip().lower() in ("a - approved", "b - approved as noted", "closed", "replied"))
        total_approved += approved
        
        pending = count - approved
        total_pending += pending
        
        row_data = [dt_name, count, approved, pending]
        bg = LIGHT if row_idx % 2 == 0 else WHITE
        for i, val in enumerate(row_data, start=2):
            c = ws.cell(row=row_idx, column=i, value=val)
            c.font = Font(color="1E2A3A")
            c.fill = fill(bg)
            c.border = thin()
            c.alignment = Alignment(horizontal="center" if i > 2 else "left")
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    row_idx += 1
    c = ws.cell(row=row_idx, column=2, value="TOTALS")
    c.font = Font(bold=True, color=PRIMARY)
    c.border = thin()
    for i, val in enumerate([total_all, total_approved, total_pending], start=3):
        c = ws.cell(row=row_idx, column=i, value=val)
        c.font = Font(bold=True, color=PRIMARY)
        c.border = thin()
        c.alignment = Alignment(horizontal="center")
        c.fill = fill("E0E7FF")

def _build_pr_register_excel(proj, dt, records, pr_items_map, pr_details_key):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PR Summary"
    raw_ws = wb.create_sheet("PR Items Raw")
    ws.sheet_view.showGridLines = False
    raw_ws.sheet_view.showGridLines = False

    PRIMARY = "1F4E78"
    ACCENT = "2563A8"
    LIGHT = "F0F4F8"
    SECTION = "E8EEF6"
    LABEL = "E2E8F0"
    SUBTLE = "F8FAFC"
    WHITE = "FFFFFF"
    MUTED = "64748B"
    TEXT = "1E2A3A"

    thin_side = Side(style="thin", color="B2B2B2")
    thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    top_rule = Border(top=thin_side)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def style_cell(cell, *, font=None, fill_color=None, border=None, align=None):
        if font: cell.font = font
        if fill_color: cell.fill = fill(fill_color)
        if border: cell.border = border
        if align: cell.alignment = align

    for col, width in {"A": 6, "B": 22, "C": 13, "D": 18, "E": 64, "F": 14}.items():
        ws.column_dimensions[col].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = False
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.left.text = "Generated from DCR System"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "PURCHASE REQUISITION REGISTER"
    style_cell(
        c,
        font=Font(name="Arial", size=16, bold=True, color=PRIMARY),
        fill_color=LIGHT,
        border=thin,
        align=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Document Control Register"
    style_cell(
        c,
        font=Font(name="Arial", size=10, italic=True, color=MUTED),
        fill_color=LIGHT,
        border=thin,
        align=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = ""
    c.fill = fill(WHITE)
    c.border = Border(bottom=thin_side)
    ws.row_dimensions[3].height = 6

    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = "PROJECT / PR INFORMATION"
    style_cell(
        c,
        font=Font(name="Arial", size=11, bold=True, color=WHITE),
        fill_color=ACCENT,
        border=thin,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[4].height = 20

    meta_fields = [
        ("Project Name", proj.get("name", "")),
        ("Project Code", proj.get("code", "")),
        ("Register", str(dt.get("name", "") or dt.get("id", "PR")).strip()),
        ("Exported", datetime.datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Total PRs", str(len(records))),
    ]
    meta_fields = [(k, v) for k, v in meta_fields if str(v or "").strip()]
    row_no = 5
    for i in range(0, len(meta_fields), 2):
        left = meta_fields[i]
        right = meta_fields[i + 1] if i + 1 < len(meta_fields) else None
        ws[f"A{row_no}"] = left[0]
        style_cell(
            ws[f"A{row_no}"],
            font=Font(name="Arial", size=10, bold=True, color=PRIMARY),
            fill_color=LABEL,
            border=thin,
            align=Alignment(vertical="center"),
        )
        ws.merge_cells(f"B{row_no}:C{row_no}")
        ws[f"B{row_no}"] = left[1]
        style_cell(
            ws[f"B{row_no}"],
            font=Font(name="Arial", size=10, color=TEXT),
            fill_color=WHITE,
            border=thin,
            align=Alignment(vertical="center", wrap_text=True),
        )
        for cell in (f"C{row_no}",):
            ws[cell].border = thin
        if right:
            ws[f"D{row_no}"] = right[0]
            style_cell(
                ws[f"D{row_no}"],
                font=Font(name="Arial", size=10, bold=True, color=PRIMARY),
                fill_color=LABEL,
                border=thin,
                align=Alignment(vertical="center"),
            )
            ws[f"E{row_no}"] = right[1]
            style_cell(
                ws[f"E{row_no}"],
                font=Font(name="Arial", size=10, color=TEXT),
                fill_color=WHITE,
                border=thin,
                align=Alignment(vertical="center", wrap_text=True),
            )
        else:
            ws[f"D{row_no}"].fill = fill(WHITE)
            ws[f"D{row_no}"].border = thin
            ws[f"E{row_no}"].fill = fill(WHITE)
            ws[f"E{row_no}"].border = thin
        ws.row_dimensions[row_no].height = 20
        row_no += 1

    ws.merge_cells(f"A{row_no}:F{row_no}")
    c = ws[f"A{row_no}"]
    c.value = "REQUISITION SUMMARY"
    style_cell(
        c,
        font=Font(name="Arial", size=11, bold=True, color=WHITE),
        fill_color=PRIMARY,
        border=thin,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[row_no].height = 20
    row_no += 1

    summary_text = (
        f"Total requisitions: {len(records)}\n"
        f"Project: {proj.get('name', '')} ({proj.get('code', '')})\n"
        f"Generated from DCR System on {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    )
    detail_lines = max(3, min(8, len(str(summary_text).splitlines()) + 1))
    ws.merge_cells(f"A{row_no}:F{row_no}")
    c = ws[f"A{row_no}"]
    c.value = summary_text
    style_cell(
        c,
        font=Font(name="Arial", size=10, color=TEXT),
        fill_color=SUBTLE,
        border=thin,
        align=Alignment(vertical="top", wrap_text=True),
    )
    ws.row_dimensions[row_no].height = detail_lines * 16
    row_no += 2

    table_header_row = row_no
    headers = ["No.", "PR Number", "PR Date", "Discipline / Trade", "PR Title", "Prepared By"]
    for idx, label in enumerate(headers, start=1):
        c = ws.cell(row=table_header_row, column=idx, value=label)
        style_cell(
            c,
            font=Font(name="Arial", size=10, bold=True, color=WHITE),
            fill_color=PRIMARY,
            border=thin,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
    ws.row_dimensions[table_header_row].height = 22
    ws.freeze_panes = f"A{table_header_row + 1}"
    ws.print_title_rows = f"{table_header_row}:{table_header_row}"
    ws.sheet_view.zoomScale = 90

    item_no = 1
    data_row = table_header_row + 1
    if records:
        for row in records:
            pr_number = _pick_first(row, ("docNo", "prNo", "prNumber")) or f"PR-{str(row.get('_id',''))[:8]}"
            pr_date = format_date(_pick_first(row, ("issuedDate", "prDate", "date")))
            discipline = _pick_first(row, ("discipline",))
            trade = _pick_first(row, ("trade",))
            disc_trade = " / ".join([v for v in [discipline, trade] if v]) or "Unspecified"
            pr_title = _pick_first(row, ("title", "subject", "description", "docTitle")) or _resolve_pr_details_value(row, pr_items_map, pr_details_key)
            prepared_by = _pick_first(row, ("preparedBy", "prepared_by", "requestedBy", "requester", "requested_by"))
            values = [item_no, pr_number, pr_date, disc_trade, pr_title, prepared_by]
            row_fill = WHITE if item_no % 2 else SUBTLE
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=col, value=val)
                style_cell(
                    cell,
                    font=Font(name="Arial", size=10, color=TEXT),
                    fill_color=row_fill,
                    border=thin,
                    align=Alignment(
                        horizontal="center" if col in (1, 3) else "left",
                        vertical="top",
                        wrap_text=True,
                    ),
                )
            ws.row_dimensions[data_row].height = 34 if len(str(pr_title or "")) > 80 else 22
            item_no += 1
            data_row += 1
    else:
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=6)
        c = ws.cell(row=data_row, column=1, value="No requisitions in this register")
        c.font = Font(name="Arial", size=10, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
        ws.row_dimensions[data_row].height = 20
        data_row += 1

    sig_row = data_row + 2
    signatures = [("Prepared By", "A", "B"), ("Reviewed By", "C", "D"), ("Approved By", "E", "F")]
    ws.merge_cells(f"A{sig_row-1}:F{sig_row-1}")
    c = ws[f"A{sig_row-1}"]
    c.value = "SIGNATURES"
    style_cell(
        c,
        font=Font(name="Arial", size=11, bold=True, color=WHITE),
        fill_color=ACCENT,
        border=thin,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[sig_row-1].height = 20
    for label, start_col, end_col in signatures:
        ws.merge_cells(f"{start_col}{sig_row}:{end_col}{sig_row}")
        top = ws[f"{start_col}{sig_row}"]
        top.value = ""
        top.border = top_rule
        top.fill = fill(WHITE)
        top.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{start_col}{sig_row+1}:{end_col}{sig_row+1}")
        lbl = ws[f"{start_col}{sig_row+1}"]
        lbl.value = label
        lbl.font = Font(name="Arial", size=9, bold=True, color=MUTED)
        lbl.alignment = Alignment(horizontal="center")
        lbl.fill = fill(SUBTLE)
    ws.row_dimensions[sig_row].height = 22
    ws.row_dimensions[sig_row + 1].height = 18

    footer_row = sig_row + 3
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = f"Generated from DCR System | Export date: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    c.font = Font(name="Arial", size=8, italic=True, color=MUTED)
    c.alignment = Alignment(horizontal="right")
    c.border = Border(top=thin_side)

    raw_headers = ["sort_order", "row_type", "description", "unit", "qty", "remarks"]
    raw_ws.merge_cells("A1:F1")
    c = raw_ws["A1"]
    c.value = "PR ITEMS RAW"
    style_cell(
        c,
        font=Font(name="Arial", size=12, bold=True, color=PRIMARY),
        fill_color=LIGHT,
        border=thin,
        align=Alignment(horizontal="center", vertical="center"),
    )
    raw_ws.row_dimensions[1].height = 22
    for idx, label in enumerate(raw_headers, start=1):
        c = raw_ws.cell(row=2, column=idx, value=label)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = fill(LABEL)
        c.border = thin
        c.alignment = Alignment(horizontal="center")
    for col, width in {1: 12, 2: 12, 3: 52, 4: 12, 5: 10, 6: 24}.items():
        raw_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    raw_ws.freeze_panes = "A3"
    raw_row = 3
    for row in records:
        pr_number = _pick_first(row, ("docNo", "prNo", "prNumber")) or f"PR-{str(row.get('_id',''))[:8]}"
        items = pr_items_map.get(row.get("_id"), [])
        raw_ws.merge_cells(start_row=raw_row, start_column=1, end_row=raw_row, end_column=6)
        c = raw_ws.cell(row=raw_row, column=1, value=pr_number)
        c.border = thin
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.font = Font(name="Arial", size=10, bold=True, color=PRIMARY)
        c.fill = fill(LABEL)
        for col in range(2, 7):
            raw_ws.cell(row=raw_row, column=col).border = thin
            raw_ws.cell(row=raw_row, column=col).fill = fill(LABEL)
        raw_ws.row_dimensions[raw_row].height = 20
        raw_row += 1
        if not items:
            values = ["", "item", "No PR items", "", "", ""]
            for col, val in enumerate(values, start=1):
                c = raw_ws.cell(row=raw_row, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(vertical="top", wrap_text=True)
            raw_row += 1
        else:
            item_sort = 1
            for it in items:
                row_type = str(it.get("row_type", "item") or "item").strip().lower()
                is_header = row_type == "header"
                values = [
                    "" if is_header else item_sort,
                    row_type,
                    str(it.get("item_name", "") or "").strip(),
                    "" if is_header else str(it.get("unit", "") or "").strip(),
                    "" if is_header else it.get("quantity", ""),
                    "" if is_header else str(it.get("remarks", "") or "").strip(),
                ]
                for col, val in enumerate(values, start=1):
                    c = raw_ws.cell(row=raw_row, column=col, value=val)
                    c.border = thin
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    if is_header:
                        c.font = Font(name="Arial", size=10, bold=True, color=PRIMARY)
                        c.fill = fill(SECTION)
                if is_header:
                    raw_ws.row_dimensions[raw_row].height = 22
                else:
                    item_sort += 1
                raw_row += 1

    if raw_row > 3:
        raw_ws.auto_filter.ref = f"A2:F{raw_row-1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@exporting_bp.route("/api/export_all/<pid>")
def api_export_all(pid):
    try:
        u = current_user()
        if u: db.log_action(u["username"],"EXPORT_EXCEL",pid,detail="Export All")
        import openpyxl

        proj    = db.get_project(pid) or {}
        dts     = db.get_doc_types(pid)
        wb      = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        PRIMARY="1F4E78"; PL="2563A8"; WHITE="FFFFFF"; ALT="F8FAFC"; OV="FFF5F5"; MUTED="9CA3AF"
        STATUS_XL = {
            "A - Approved":              ("C6EFCE","064E3B"),
            "B - Approved As Noted":     ("C6EFCE","064E3B"),
            "B,C - Approved & Resubmit": ("FFEB9C","713F12"),
            "C - Revise & Resubmit":     ("FFC7CE","991B1B"),
            "D - Review not Required":   ("FFC7CE","991B1B"),
            "Under Review":              ("FFEB9C","713F12"),
            "Cancelled":                 ("FFC7CE","991B1B"),
            "Open":                      ("FFEB9C","713F12"),
            "Closed":                    ("C6EFCE","064E3B"),
            "Replied":                   ("C6EFCE","064E3B"),
            "Pending":                   ("FFEB9C","713F12"),
        }
        def fill(c): return PatternFill("solid", fgColor=c)
        def thin(): s=Side(style="thin",color="B2B2B2"); return Border(left=s,right=s,top=s,bottom=s)

        from flask import request
        days = request.args.get('days')
        status_filter = request.args.get('status')
        
        cutoff = None
        if days and days != 'all':
            from datetime import datetime, timedelta
            try: cutoff = (datetime.now() - timedelta(days=int(days))).strftime("%Y-%m-%d")
            except: pass

        records_by_dt = {}
        for dt in dts:
            cols    = [c for c in db.get_columns(pid, dt["id"]) if c["visible"]]
            records = sorted(db.get_records(pid, dt["id"]), key=_doc_revision_sort_key)
            
            # Apply Filters
            if cutoff:
                records = [r for r in records if r.get('issuedDate', '') >= cutoff or r.get('receivedDate', '') >= cutoff]
            if status_filter == 'overdue':
                from utils import is_overdue
                expected_reply_rule = db.get_expected_reply_rule(pid, dt["id"])
                has_exp_col = any(c["col_key"]=="expectedReplyDate" for c in cols)
                records = [r for r in records if is_overdue(r.get("issuedDate"), r.get("docNo"), r.get("actualReplyDate"), has_exp_col, expected_reply_rule, status=r.get("status"), action=r.get("action"))]
                
            records_by_dt[dt["name"] or dt["id"]] = records
            
            is_pr = _is_pr_dt(dt)
            pr_details_key = _pr_details_key(cols) if is_pr else None
            pr_items_map = db.get_pr_items_for_records([r.get("_id") for r in records]) if is_pr else {}
            web_widths = db.get_col_widths(pid, dt["id"])
            ws = wb.create_sheet(title=dt["id"][:31])
            _write_register_excel_sheet(ws, proj, dt, cols, records, pr_items_map, pr_details_key, web_widths)

        if not wb.sheetnames:
            ws = wb.create_sheet("Empty"); ws.cell(1,1,"No data")
            
        # Build Summary Dashboard
        summary_ws = wb.create_sheet(title="Dashboard", index=0)
        _write_summary_dashboard(summary_ws, proj, records_by_dt)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"{proj.get('code','DCR')}_All_Registers.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return traceback.format_exc(), 500


@exporting_bp.route("/api/export/<pid>/<dt_id>")
def api_export(pid, dt_id):
    try:
        import openpyxl
        from flask import request

        proj    = db.get_project(pid) or {}
        cols    = [c for c in db.get_columns(pid, dt_id) if c["visible"]]
        records = sorted(db.get_records(pid, dt_id), key=_doc_revision_sort_key)
        
        # Apply Filters
        days = request.args.get('days')
        status_filter = request.args.get('status')
        if days and days != 'all':
            from datetime import datetime, timedelta
            try:
                cutoff = (datetime.now() - timedelta(days=int(days))).strftime("%Y-%m-%d")
                records = [r for r in records if r.get('issuedDate', '') >= cutoff or r.get('receivedDate', '') >= cutoff]
            except: pass
        if status_filter == 'overdue':
            from utils import is_overdue
            expected_reply_rule = db.get_expected_reply_rule(pid, dt_id)
            has_exp_col = any(c["col_key"]=="expectedReplyDate" for c in cols)
            records = [r for r in records if is_overdue(r.get("issuedDate"), r.get("docNo"), r.get("actualReplyDate"), has_exp_col, expected_reply_rule, status=r.get("status"), action=r.get("action"))]
        dts     = db.get_doc_types(pid)
        dt      = next((d for d in dts if d["id"] == dt_id), None)
        is_pr = _is_pr_dt(dt)
        pr_details_key = _pr_details_key(cols) if is_pr else None
        pr_items_map = db.get_pr_items_for_records([r.get("_id") for r in records]) if is_pr else {}
        web_widths = db.get_col_widths(pid, dt_id)

        if is_pr:
            buf = _build_pr_register_excel(proj, dt or {"id": dt_id, "name": dt_id}, records, pr_items_map, pr_details_key)
            fname = f"{_safe_excel_name_part(proj.get('code','DCR'), 'DCR')}_{_safe_excel_name_part(dt_id, 'PR')}_Register.xlsx"
            return send_file(
                buf,
                as_attachment=True,
                download_name=fname,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = dt_id[:31]
        _write_register_excel_sheet(ws, proj, dt or {"id": dt_id, "name": dt_id}, cols, records, pr_items_map, pr_details_key, web_widths)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"{proj.get('code','DCR')}_{dt_id}_Register.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return traceback.format_exc(), 500





def _build_executive_summary_pdf(pid, dt_id=None):
    from flask import render_template

    import datetime
    
    proj = db.get_project(pid) or {}
    dts = db.get_doc_types(pid)
    
    if dt_id and dt_id != "all":
        dts = [d for d in dts if d["id"] == dt_id]
        
    registers_data = []
    total_docs = 0
    total_approved = 0
    total_pending = 0
    
    # Check date filters
    from flask import request
    days = request.args.get('days')
    status_filter = request.args.get('status')
    cutoff = None
    if days and days != 'all':
        try: cutoff = (datetime.datetime.now() - datetime.timedelta(days=int(days))).strftime("%Y-%m-%d")
        except: pass

    for dt in dts:
        records = db.get_records(pid, dt["id"])
        
        # Apply Filters
        if cutoff:
            records = [r for r in records if r.get('issuedDate', '') >= cutoff or r.get('receivedDate', '') >= cutoff]
        if status_filter == 'overdue':
            from utils import is_overdue
            expected_reply_rule = db.get_expected_reply_rule(pid, dt["id"])
            cols = db.get_columns(pid, dt["id"])
            has_exp_col = any(c["col_key"]=="expectedReplyDate" for c in cols)
            records = [r for r in records if is_overdue(r.get("issuedDate"), r.get("docNo"), r.get("actualReplyDate"), has_exp_col, expected_reply_rule, status=r.get("status"), action=r.get("action"))]

        count = len(records)
        if count == 0:
            registers_data.append({"name": dt.get("name") or dt.get("id"), "total": 0, "approved": 0, "pending": 0})
            continue
            
        approved = sum(1 for r in records if str(r.get("status", "")).strip().lower() in ("a - approved", "b - approved as noted", "closed", "replied"))
        pending = count - approved
        
        total_docs += count
        total_approved += approved
        total_pending += pending
        
        registers_data.append({
            "name": dt.get("name") or dt.get("id"),
            "total": count,
            "approved": approved,
            "pending": pending,
            "records": records,
            "cols": db.get_columns(pid, dt["id"])
        })
        
    logo_l = db.get_logo(pid, "logo_left")
    logo_r = db.get_logo(pid, "logo_right")
    if logo_l and "," in logo_l: logo_l = logo_l.split(",", 1)[1]
    if logo_r and "," in logo_r: logo_r = logo_r.split(",", 1)[1]

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    import io
    from PIL import Image as PILImage
    import base64
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=5,
        alignment=1 # Center
    )
    sub_style = ParagraphStyle(
        'SubStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor("#475569"),
        spaceAfter=20,
        alignment=1
    )
    
    # Title
    elements.append(Paragraph("EXECUTIVE SUMMARY REPORT", title_style))
    elements.append(Paragraph(f"{proj.get('name', 'Project')} ({proj.get('code', 'DCR')})", sub_style))
    
    # Meta Box
    meta_data = [
        ["Generated On", "Total Documents", "Completion Status"],
        [
            datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
            str(total_docs),
            f"{total_approved / total_docs * 100:.1f}%" if total_docs > 0 else "0%"
        ]
    ]
    meta_table = Table(meta_data, colWidths=[60*mm, 60*mm, 60*mm])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#64748b")),
        ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor("#0f172a")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 2),
        ('TOPPADDING', (0,1), (-1,1), 2),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#cbd5e1")),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 20*mm))
    
    # Registers Table
    elements.append(Paragraph("Document Registers Summary", styles['Heading2']))
    table_data = [["Document Type", "Total Records", "Approved / Closed", "Pending / Overdue", "Status"]]
    
    for dt in registers_data:
        status_text = "Up to date" if dt["pending"] == 0 and dt["total"] > 0 else ("Action Required" if dt["pending"] > 0 else "Empty")
        table_data.append([
            dt["name"],
            str(dt["total"]),
            str(dt["approved"]),
            str(dt["pending"]),
            status_text
        ])
    
    # Grand Total
    table_data.append([
        "GRAND TOTAL",
        str(total_docs),
        str(total_approved),
        str(total_pending),
        ""
    ])
    
    reg_table = Table(table_data, colWidths=[55*mm, 30*mm, 35*mm, 35*mm, 25*mm])
    reg_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e40af")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('INNERGRID', (0,0), (-1,-2), 0.25, colors.HexColor("#e2e8f0")),
        ('BOX', (0,0), (-1,-2), 0.25, colors.HexColor("#e2e8f0")),
    ]
    
    # Zebra striping
    for i in range(1, len(table_data)-1):
        if i % 2 == 0:
            reg_style.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor("#f8fafc")))
            
    # Highlight columns
    for i in range(1, len(table_data)):
        reg_style.append(('TEXTCOLOR', (2,i), (2,i), colors.HexColor("#166534"))) # Green approved
        reg_style.append(('TEXTCOLOR', (3,i), (3,i), colors.HexColor("#991b1b"))) # Red pending
        reg_style.append(('FONTNAME', (2,i), (3,i), 'Helvetica-Bold'))
        
    # Grand Total row styling
    reg_style.extend([
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#eff6ff")),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor("#1e40af")),
    ])
    
    reg_table.setStyle(TableStyle(reg_style))
    elements.append(reg_table)
    
    # ---------------------------------------------------------
    # Generate the missing pages: A separate table for each DT
    # ---------------------------------------------------------
    from reportlab.platypus import PageBreak
    
    for dt_data in registers_data:
        records = dt_data.get("records", [])
        if not records: continue
        
        elements.append(PageBreak())
        elements.append(Paragraph(f"{dt_data['name']} - Document Register", styles['Heading2']))
        elements.append(Spacer(1, 10*mm))
        
        cols = [c for c in dt_data.get("cols", []) if c.get("visible")]
        if not cols: continue
        
        header_row = ["No."] + [c["label"] for c in cols]
        dt_table_data = [header_row]
        
        for idx, r in enumerate(records, 1):
            row_data = [str(idx)]
            for c in cols:
                key = c["col_key"]
                val = str(r.get(key, "") or "")
                # Limit length so PDF table cell doesn't break vertically if it's too huge
                if len(val) > 100: val = val[:97] + "..."
                row_data.append(val)
            dt_table_data.append(row_data)
            
        # Calculate proportional widths based on 170mm total available width
        w_per_col = (170 / len(cols)) * mm if len(cols) > 0 else 10*mm
        dt_col_widths = [12*mm] + [w_per_col] * len(cols)
        
        dt_table = Table(dt_table_data, colWidths=dt_col_widths, repeatRows=1)
        dt_table_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e40af")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.HexColor("#e2e8f0")),
            ('BOX', (0,0), (-1,-1), 0.25, colors.HexColor("#e2e8f0")),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]
        
        # Zebra striping
        for i in range(1, len(dt_table_data)):
            if i % 2 == 0:
                dt_table_style.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor("#f8fafc")))
                
        dt_table.setStyle(TableStyle(dt_table_style))
        elements.append(dt_table)
    
    # Build PDF
    doc.build(elements)
    buf.seek(0)
    return buf

@exporting_bp.route("/api/export_pdf/<pid>/<dt_id>")
def api_export_pdf(pid, dt_id):
    try:

        buf = _build_executive_summary_pdf(pid, dt_id)
        fname = f"{(db.get_project(pid) or {}).get('code','DCR')}_{dt_id}_Executive_Summary.pdf"
        return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")
    except Exception as e:
        import traceback
        return traceback.format_exc(), 500


@exporting_bp.route("/api/export_pdf_all/<pid>")
def api_export_pdf_all(pid):
    try:

        buf = _build_executive_summary_pdf(pid, "all")
        fname = f"{(db.get_project(pid) or {}).get('code','DCR')}_Executive_Summary.pdf"
        return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")
    except Exception as e:
        import traceback
        return traceback.format_exc(), 500


@exporting_bp.route("/api/import/<pid>/<dt_id>", methods=["POST"])
def api_import(pid, dt_id):
    if not can_edit(pid): return jsonify(error="LOGIN_REQUIRED"), 403
    data = request.get_json(silent=True) or {}
    b64  = data.get("file_b64","")
    ext  = data.get("ext","csv")
    if "," in b64: b64 = b64.split(",",1)[1]

    cols    = db.get_columns(pid, dt_id)
    col_map = {c["label"]: c["col_key"] for c in cols}
    imported = 0
    created = 0
    updated = 0
    skipped_blank = 0
    skipped_invalid = 0
    warnings = []

    try:
        if ext in ("xlsx","xls"):
            import openpyxl
            wb  = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
            ws  = wb.active
            imported, created, updated, _, skipped_blank, skipped_invalid, warnings = _import_excel_worksheet(pid, dt_id, ws, cols)
        else:
            import csv, datetime as _dt
            text   = base64.b64decode(b64).decode("utf-8","ignore")
            header = None
            date_cols = {c["col_key"] for c in cols if c.get("col_type") in ("date","auto_date")}
            for row_idx, line in enumerate(csv.reader(io.StringIO(text)), start=1):
                if not header:
                    if any(c in line for c in ["Document No.","Sr.","docNo"]):
                        header = [col_map.get(h.strip(),h.strip()) for h in line]
                    continue
                if not any(line):
                    skipped_blank += 1
                    continue
                row_data = {}
                for i,val in enumerate(line):
                    if i<len(header) and header[i] and header[i] not in ("sr","Sr.","Sr",""):
                        v = val.strip()
                        if header[i] in date_cols and v:
                            for fmt in ("%d/%b/%Y","%d-%b-%Y","%Y-%m-%d","%d/%m/%Y"):
                                try: v = _dt.datetime.strptime(v,fmt).strftime("%Y-%m-%d"); break
                                except: pass
                        row_data[header[i]] = v
                if not _has_meaningful_values(row_data):
                    skipped_blank += 1
                    continue
                try:
                    action = _save_import_row(pid, dt_id, row_data)
                    imported += 1
                    if action == "updated":
                        updated += 1
                    else:
                        created += 1
                except Exception as e:
                    skipped_invalid += 1
                    logger.warning("import_csv_row_skipped pid=%s dt_id=%s row=%s error=%s",
                                   pid, dt_id, row_idx, e)
                    if len(warnings) < 20:
                        warnings.append({"row": row_idx, "error": str(e)})
    except Exception as e:
        logger.error("import_failed pid=%s dt_id=%s ext=%s error=%s", pid, dt_id, ext, e)
        return jsonify(ok=False, error=str(e)), 500

    return jsonify(
        ok=True,
        imported=imported,
        created=created,
        updated=updated,
        skipped_blank=skipped_blank,
        skipped_invalid=skipped_invalid,
        warnings=warnings,
    )


@exporting_bp.route("/api/import_project/<pid>", methods=["POST"])
def api_import_project(pid):
    if not can_edit(pid): return jsonify(error="LOGIN_REQUIRED"), 403
    data = request.get_json(silent=True) or {}
    b64  = data.get("file_b64","")
    ext  = data.get("ext","xlsx")
    if "," in b64: b64 = b64.split(",",1)[1]
    if ext not in ("xlsx","xls"):
        return jsonify(ok=False, error="Only Excel workbooks are supported"), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True)
        dts = db.get_doc_types(pid)
        imported_total = 0
        created_total = 0
        updated_total = 0
        skipped_blank_total = 0
        skipped_invalid_total = 0
        matched_sheets = []
        skipped_sheets = []
        errors = []
        warnings = []

        for ws in wb.worksheets:
            dt = _match_sheet_to_dt(ws.title, dts)
            if not dt:
                logger.warning("import_project_sheet_skipped pid=%s sheet=%s reason=no_matching_document_type",
                               pid, ws.title)
                skipped_sheets.append({"sheet": ws.title, "reason": "No matching document type"})
                continue
            try:
                cols = db.get_columns(pid, dt["id"])
                imported, created, updated, has_header, skipped_blank, skipped_invalid, sheet_warnings = _import_excel_worksheet(pid, dt["id"], ws, cols)
                if not has_header:
                    logger.warning("import_project_sheet_skipped pid=%s dt_id=%s sheet=%s reason=no_valid_header",
                                   pid, dt["id"], ws.title)
                    skipped_sheets.append({"sheet": ws.title, "reason": "No valid header row"})
                    continue
                imported_total += imported
                created_total += created
                updated_total += updated
                skipped_blank_total += skipped_blank
                skipped_invalid_total += skipped_invalid
                matched_sheets.append({
                    "sheet": ws.title,
                    "dt_id": dt["id"],
                    "dt_name": dt["name"],
                    "imported": imported,
                    "created": created,
                    "updated": updated,
                    "skipped_blank": skipped_blank,
                    "skipped_invalid": skipped_invalid,
                })
                if sheet_warnings and len(warnings) < 20:
                    for w in sheet_warnings:
                        warnings.append({"sheet": ws.title, **w})
                        if len(warnings) >= 20:
                            break
            except Exception as e:
                logger.error("import_project_sheet_failed pid=%s dt_id=%s sheet=%s error=%s",
                             pid, dt.get("id",""), ws.title, e)
                errors.append({"sheet": ws.title, "error": str(e)})

        return jsonify(
            ok=True,
            project_id=pid,
            imported_total=imported_total,
            created_total=created_total,
            updated_total=updated_total,
            skipped_blank_total=skipped_blank_total,
            skipped_invalid_total=skipped_invalid_total,
            matched_sheets=matched_sheets,
            skipped_sheets=skipped_sheets,
            errors=errors,
            warnings=warnings,
        )
    except Exception as e:
        logger.error("import_project_failed pid=%s ext=%s error=%s", pid, ext, e)
        return jsonify(ok=False, error=str(e)), 500


